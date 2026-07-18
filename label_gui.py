import sys
import os
import json
import re
import shutil
import site
import threading
import zipfile
from datetime import datetime

if sys.stdout is not None:
    sys.stdout.reconfigure(encoding="utf-8")
if sys.stderr is not None:
    sys.stderr.reconfigure(encoding="utf-8")

# Harmless fallback for running this as a plain folder (python label_gui.py)
# on a machine without Pillow/pywin32 installed - a `vendor` folder here, if
# present, gets picked up automatically. Not needed for the packaged .exe
# build (PyInstaller bundles these directly); addsitedir on a missing folder
# is a silent no-op either way.
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
site.addsitedir(os.path.join(SCRIPT_DIR, "vendor"))

import tkinter as tk
from tkinter import messagebox, simpledialog, filedialog, ttk

from PIL import Image, ImageDraw, ImageFont, ImageWin, ImageTk

import storage
import app_settings
import local_server
import ai_assist
import knowledge

FAVORITES_PATH = os.path.join(storage.APP_DATA_DIR, "favorites.json")
DEBUG_PREVIEW_PATH = os.path.join(storage.APP_DATA_DIR, "_last_label_preview.png")


def load_favorites():
    if not os.path.isfile(FAVORITES_PATH):
        return {}
    try:
        with open(FAVORITES_PATH, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return {}


def save_favorites(favorites):
    os.makedirs(storage.APP_DATA_DIR, exist_ok=True)
    with open(FAVORITES_PATH, "w", encoding="utf-8") as f:
        json.dump(favorites, f, ensure_ascii=False, indent=2)


APP_VERSION = "1.17.0"

DOTS_PER_MM = 8  # matches standard 203dpi thermal label printers

TIME_OPTIONS = storage.TIME_OPTIONS  # เช้า/เที่ยง/เย็น/ก่อนนอน
UNIT_OPTIONS = ["เม็ด", "แคปซูล", "ช้อนชา", "ช้อนโต๊ะ", "ซอง", "ml", "หยด", "พ่น"]

# Per-label language - chosen at print time (row dropdown, default "th"),
# not a per-drug/global setting, since a shop's customers are often mixed
# nationality and the right language varies visit to visit. Only the FIXED
# template wording (field labels, dose phrasing, warning line) is
# translated here - drug names/notes/extra_labels are whatever the
# pharmacist actually typed and are never auto-translated (medical-accuracy
# risk), so English/Burmese output still needs those typed in that language.
LABEL_LANGS = ["th", "en", "mm"]
LABEL_LANG_NAMES = {"th": "ไทย", "en": "English", "mm": "မြန်မာ"}
TIME_EN = {"เช้า": "Morning", "เที่ยง": "Noon", "เย็น": "Evening", "ก่อนนอน": "Before bed"}
MEAL_EN = {"ก่อนอาหาร": "Before meals", "หลังอาหาร": "After meals"}
UNIT_EN = {
    "เม็ด": "tab(s)", "แคปซูล": "cap(s)", "ช้อนชา": "tsp(5ml)", "ช้อนโต๊ะ": "tbsp(15ml)",
    "ซอง": "sachet(s)", "ml": "ml", "หยด": "drop(s)", "พ่น": "spray(s)",
}
# Burmese - translated/reviewed by a Burmese-speaking pharmacist contact
# (not machine-translated), same caution as the English set above.
TIME_MM = {"เช้า": "မနက်", "เที่ยง": "နေ့ခင်း", "เย็น": "ညနေ", "ก่อนนอน": "အိပ်ခါနီး"}
MEAL_MM = {"ก่อนอาหาร": "အစားမစားမီ", "หลังอาหาร": "အစားစားပြီး"}
UNIT_MM = {
    "เม็ด": "ဆေးပြား", "แคปซูล": "ဆေးတောင့်", "ช้อนชา": "လက်ဖက်ရည်ဇွန်း", "ช้อนโต๊ะ": "စားပွဲတင်ဇွန်း",
    "ซอง": "ထုပ်", "ml": "ml", "หยด": "စက်", "พ่น": "ဖျန်း",
}
TIME_TR = {"en": TIME_EN, "mm": TIME_MM}
MEAL_TR = {"en": MEAL_EN, "mm": MEAL_MM}
UNIT_TR = {"en": UNIT_EN, "mm": UNIT_MM}


def _tr_times(times, lang):
    table = TIME_TR.get(lang)
    if not table:
        return list(times or [])
    return [table.get(t, t) for t in (times or [])]


def _tr_meal(meal, lang):
    table = MEAL_TR.get(lang)
    if not table:
        return meal or ""
    return table.get(meal, meal or "")


def _tr_unit(unit, lang):
    table = UNIT_TR.get(lang)
    if not table:
        return unit or "เม็ด"
    return table.get(unit, unit or "เม็ด")

EXTRA_LABEL_OPTIONS_BY_MODE = {
    "oral": [
        "ทานยาก่อนอาหาร 1/2-1 ชม",
        "ทานยาหลังอาหารทันที",
        "ทานติดต่อกันจนหมด",
        "ดื่มน้ำตามมากๆ",
        "ยานี้อาจทำให้ง่วงซึม",
        "ห้ามรับประทานพร้อมนม ยาลดกรด",
        "ทานเมื่อมีอาการ",
    ],
    "topical": [
        "หายแล้ว ทาต่ออีก 1 สัปดาห์",
        "ทาต่อเนื่อง 2 สัปดาห์",
    ],
    "drops": [
        "หยดเมื่อมีอาการ",
        "หยดต่อเนื่อง 2 สัปดาห์",
        "ยาหยอดตาใช้ได้ 3 เดือน",
    ],
}
MAX_EXTRA_LABELS = 2
# Longest a free-typed extra label can be before it'd force the label
# renderer to shrink below a comfortable reading size - measured against the
# widest Thai glyphs at the extra-labels' starting font size (24pt bold) on
# a default 80mm-wide label, with a small safety margin.
MAX_CUSTOM_EXTRA_LABEL_CHARS = 30

USAGE_MODES = ["oral", "topical", "drops"]
USAGE_MODE_LABELS = {"oral": "ยากิน", "topical": "ยาทา", "drops": "ยาหยอด"}

MEAL_VALUE_TO_DISPLAY = {"ก่อนอาหาร": "ก่อนอาหาร", "หลังอาหาร": "หลังอาหาร", "": "ไม่ระบุก่อน/หลังอาหาร"}
MEAL_DISPLAY_TO_VALUE = {v: k for k, v in MEAL_VALUE_TO_DISPLAY.items()}
MEAL_OPTIONS_DISPLAY = list(MEAL_VALUE_TO_DISPLAY.values())

UI_SCALE = 1.3  # GUI-only (window/fonts/spacing) - does not affect the physical label image


def fs(n):
    """Scale a GUI font size or pixel dimension by UI_SCALE."""
    return round(n * UI_SCALE)


# thin wrappers so the rest of this file (ported from the DB-backed version)
# doesn't need to change its call sites
def search_products(term, limit=30):
    return storage.search_templates(term, limit=limit)


def get_product_med_info(idproduct):
    return storage.get_template(idproduct)


has_dosing_data = storage.has_dosing_data


def save_product_med_info(idproduct, drug):
    """Returns the row id (new id for a brand-new drug, same id otherwise)."""
    return storage.upsert_template(idproduct, drug)


DEFAULT_EXCEL_NAME_COLUMNS = ["ชื่อการค้า", "ชื่อสินค้า", "ชื่อยา", "ชื่อการค้า/ชื่อสินค้า", "col1"]
DEFAULT_EXCEL_BARCODE_COLUMNS = ["บาร์โค้ด", "Barcode", "barcode", "บาร์โค้ดในหน่วยนับ 1", "col2"]
DEFAULT_EXCEL_GENERIC_COLUMNS = ["ชื่อสามัญทางยา", "ชื่อสามัญ", "Generic Name", "generic name", "col3"]


def export_backup(zip_path):
    """Bundle everything needed to move to another machine into one zip:
    the whole drug database, favorites, and shop/pharmacist settings.
    printer_name is deliberately excluded - the target machine almost
    certainly has a different printer installed, and importing a printer
    name that doesn't exist there would silently break printing."""
    settings = app_settings.load_settings()
    export_settings = {k: v for k, v in settings.items() if k != "printer_name"}
    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
        if os.path.isfile(storage.DB_PATH):
            zf.write(storage.DB_PATH, arcname="data.db")
        if os.path.isfile(FAVORITES_PATH):
            zf.write(FAVORITES_PATH, arcname="favorites.json")
        zf.writestr("settings_export.json", json.dumps(export_settings, ensure_ascii=False, indent=2))


def import_backup(zip_path):
    """Restore a zip made by export_backup(). Overwrites the local drug
    database and favorites entirely; merges settings but keeps this
    machine's own printer_name untouched."""
    os.makedirs(storage.APP_DATA_DIR, exist_ok=True)
    with zipfile.ZipFile(zip_path, "r") as zf:
        names = zf.namelist()
        if "data.db" not in names:
            raise ValueError("ไฟล์นี้ไม่ใช่ไฟล์สำรองข้อมูลของโปรแกรมนี้ (ไม่พบ data.db ข้างใน)")
        with zf.open("data.db") as src, open(storage.DB_PATH, "wb") as dst:
            shutil.copyfileobj(src, dst)
        if "favorites.json" in names:
            with zf.open("favorites.json") as src, open(FAVORITES_PATH, "wb") as dst:
                shutil.copyfileobj(src, dst)
        if "settings_export.json" in names:
            imported_settings = json.loads(zf.read("settings_export.json").decode("utf-8"))
            current = app_settings.load_settings()
            current.update({k: v for k, v in imported_settings.items() if k != "printer_name"})
            app_settings.save_settings(current)


_MAX_HEADER_SCAN_ROWS = 10


def read_excel_drug_names_and_barcodes(path, name_column=None, barcode_column=None, generic_column=None):
    """Read drug trade names (plus optional barcode/generic-name columns)
    out of the first sheet of an .xlsx file - covers a fresh import (names
    + barcodes + generic names together) and adding barcodes/generic names
    to drugs that already exist (see storage.bulk_import_names_and_barcodes(),
    which treats a name match as "update barcode/generic name only, don't
    touch dosing"). Scans the first several rows (not just row 1) for header
    matches - some exported files have a title row or blank row above the
    real header, or even a genuinely two-row header where sub-columns (e.g.
    a barcode column) sit one row below the main header (a real POS export
    shape - the name/generic columns are on the "category" row, the barcode
    column is on the "sub-column" row directly under it). Each of
    name/barcode/generic is searched independently across every scanned
    header row - they don't need to share one row. Matching is
    case-insensitive and ignores extra whitespace. If a *_column arg is
    blank, tries the matching DEFAULT_EXCEL_*_COLUMNS list in order;
    barcode/generic are best-effort - if not found, every row's value for
    that field is just empty. Raises ValueError with a Thai message
    (including what headers were actually seen) only if the NAME column
    can't be found."""
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)

    name_column = (name_column or "").strip()
    name_candidates_lower = [c.lower() for c in ([name_column] if name_column else DEFAULT_EXCEL_NAME_COLUMNS)]
    barcode_column = (barcode_column or "").strip()
    barcode_candidates_lower = [c.lower() for c in ([barcode_column] if barcode_column else DEFAULT_EXCEL_BARCODE_COLUMNS)]
    generic_column = (generic_column or "").strip()
    generic_candidates_lower = [c.lower() for c in ([generic_column] if generic_column else DEFAULT_EXCEL_GENERIC_COLUMNS)]

    name_idx = barcode_idx = generic_idx = None
    last_header_row_num = 0
    scanned_headers = []
    for row_num, row in enumerate(rows, start=1):
        if row_num > _MAX_HEADER_SCAN_ROWS:
            break
        header = [str(h).strip() if h is not None else "" for h in row]
        header_lower = [h.lower() for h in header]
        scanned_headers.append((row_num, header))
        if name_idx is None:
            for cand_lower in name_candidates_lower:
                if cand_lower in header_lower:
                    name_idx = header_lower.index(cand_lower)
                    last_header_row_num = row_num
                    break
        if barcode_idx is None:
            for cand_lower in barcode_candidates_lower:
                if cand_lower in header_lower:
                    barcode_idx = header_lower.index(cand_lower)
                    last_header_row_num = max(last_header_row_num, row_num)
                    break
        if generic_idx is None:
            for cand_lower in generic_candidates_lower:
                if cand_lower in header_lower:
                    generic_idx = header_lower.index(cand_lower)
                    last_header_row_num = max(last_header_row_num, row_num)
                    break

    if name_idx is None:
        hint = (
            f"\n** ตรวจสอบว่าชื่อ column ที่พิมพ์ ตรงกับหัวตารางในไฟล์ Excel เป๊ะๆ "
            f"(เช็คให้แล้ว {_MAX_HEADER_SCAN_ROWS} แถวแรกของไฟล์ ไม่ต้องอยู่แถว 1 พอดีก็ได้) **"
        )
        seen = "; ".join(
            f"แถว {n}: {', '.join(h) if any(h) else '(ว่างเปล่า)'}" for n, h in scanned_headers
        )
        if name_column:
            raise ValueError(f"ไม่พบ column ชื่อ '{name_column}' ในไฟล์ (เช็คแล้ว {len(scanned_headers)} แถวแรก - {seen}){hint}")
        raise ValueError(
            "ไม่พบ column ชื่อ " + "/".join(DEFAULT_EXCEL_NAME_COLUMNS) +
            f" ในไฟล์ (เช็คแล้ว {len(scanned_headers)} แถวแรก - {seen}) กรุณาระบุชื่อ column เอง{hint}"
        )

    rows_out = []
    wb2 = load_workbook(path, read_only=True, data_only=True)
    ws2 = wb2.active
    for row_num, row in enumerate(ws2.iter_rows(values_only=True), start=1):
        if row_num <= last_header_row_num:
            continue
        if name_idx < len(row) and row[name_idx] is not None:
            name = str(row[name_idx])
            barcode = ""
            if barcode_idx is not None and barcode_idx < len(row) and row[barcode_idx] is not None:
                barcode = str(row[barcode_idx])
            generic = ""
            if generic_idx is not None and generic_idx < len(row) and row[generic_idx] is not None:
                generic = str(row[generic_idx])
            rows_out.append((name, barcode, generic))
    return rows_out


def find_font(size, bold=False, lang="th"):
    # Tahoma/Leelawadee UI have no Myanmar glyphs at all - a Burmese label
    # needs the Windows-bundled Myanmar Text font instead, for every string
    # (including digits/EXP), not just the translated phrases.
    if lang == "mm":
        candidates = [r"C:\Windows\Fonts\mmrtextb.ttf"] if bold else [r"C:\Windows\Fonts\mmrtext.ttf"]
    else:
        candidates = (
            [r"C:\Windows\Fonts\tahomabd.ttf", r"C:\Windows\Fonts\leelawui.ttf"]
            if bold else
            [r"C:\Windows\Fonts\tahoma.ttf", r"C:\Windows\Fonts\leelawue.ttf"]
        )
    for path in candidates:
        if os.path.isfile(path):
            return ImageFont.truetype(path, size)
    return ImageFont.load_default()


def detect_font_lang(text):
    """Free-typed fields (patient name, drug1/drug2, note) can hold Thai,
    English, or Burmese text depending on what the pharmacist actually
    typed there - pick the font by inspecting the real characters instead
    of trusting the label-language dropdown, since e.g. a patient's name
    is Thai even on a Burmese-language label, while a pharmacist who has
    an expert's Burmese translation for the indication should have that
    render correctly too."""
    for ch in text or "":
        if 0x1000 <= ord(ch) <= 0x109F:
            return "mm"
    return "th"


def fit_font(draw, text, max_width, start_size, bold=False, min_size=14, lang="th"):
    """Shrink font size until text fits max_width, so long drug/patient
    names never overflow past the edge of the sticker."""
    size = start_size
    font = find_font(size, bold=bold, lang=lang)
    while size > min_size and draw.textlength(text, font=font) > max_width:
        size -= 1
        font = find_font(size, bold=bold, lang=lang)
    return font


# This Pillow install has no libraqm (no Windows wheel ships it), so it can't
# shape stacked Thai combining marks. Any base char that carries BOTH
# mai-han-akat (ั) and a tone mark (่ ้ ๊ ๋) - e.g. "นั่น", "ทั้ง", "มั่นใจ",
# "สั่ง" - draws both marks at the exact same font-designed anchor point and
# the tone mark disappears underneath mai-han-akat. Fix: draw the base+ั
# normally, then draw the tone mark by itself shifted up so it stacks above
# instead of overlapping. Verified against สนั่น/ทั้งหมด/มั่นใจ/สั่งซื้อ/ทั่วไป
# while leaving single-mark text (ก็/น้ำ/ชื่อ) unaffected.
_MAI_HAN_AKAT = "ั"
_TONE_MARKS = "่้๊๋"


def draw_thai_text(draw, xy, text, font, fill=0):
    x, y = xy
    text = text or ""
    i = 0
    n = len(text)
    while i < n:
        ch = text[i]
        if (
            i + 2 < n + 1 and i + 1 < n and text[i + 1] == _MAI_HAN_AKAT
            and i + 2 < n and text[i + 2] in _TONE_MARKS
        ):
            base = ch
            tone = text[i + 2]
            cluster = base + _MAI_HAN_AKAT
            draw.text((x, y), cluster, font=font, fill=fill)
            base_w = draw.textlength(base, font=font)
            lift = int(font.size * 0.32)
            draw.text((x + base_w, y - lift), tone, font=font, fill=fill)
            x += draw.textlength(cluster, font=font)
            i += 3
            continue
        draw.text((x, y), ch, font=font, fill=fill)
        x += draw.textlength(ch, font=font)
        i += 1


def ask_upload_note(parent):
    """Bigger note-entry dialog for patient document uploads (replaces
    simpledialog.askstring, whose fixed small entry/font was hard to use
    for longer notes). Returns the note string, or None if cancelled."""
    win = tk.Toplevel(parent)
    win.title("หมายเหตุ")
    win.transient(parent)
    win.grab_set()

    result = {"value": None}

    tk.Label(
        win, text="ใส่หมายเหตุ (ถ้ามี) แล้วกด OK เพื่ออัปโหลด:",
        font=("Tahoma", fs(20)), anchor="w",
    ).pack(fill="x", padx=fs(12), pady=(fs(12), fs(6)))

    text_box = tk.Text(win, font=("Tahoma", fs(20)), width=45, height=9, wrap="word")
    text_box.pack(fill="both", expand=True, padx=fs(12), pady=(0, fs(10)))
    text_box.focus_set()

    def on_ok(event=None):
        result["value"] = text_box.get("1.0", "end-1c").strip()
        win.destroy()

    def on_cancel(event=None):
        result["value"] = None
        win.destroy()

    btn_row = tk.Frame(win)
    btn_row.pack(fill="x", padx=fs(12), pady=(0, fs(12)))
    tk.Button(btn_row, text="OK", font=("Tahoma", fs(11)), width=10, command=on_ok).pack(side="right", padx=(fs(6), 0))
    tk.Button(btn_row, text="Cancel", font=("Tahoma", fs(11)), width=10, command=on_cancel).pack(side="right")

    win.protocol("WM_DELETE_WINDOW", on_cancel)
    win.bind("<Escape>", on_cancel)
    win.update_idletasks()
    win.minsize(win.winfo_width(), win.winfo_height())
    win.wait_window()
    return result["value"]


def compute_dose_lines(data):
    """Returns (dose_text, line2) - the same wording used on the physical
    label. Shared between build_label_image() (rendering) and the print
    history dialog (showing dosing info on click) so the two can never
    drift apart. data["lang"] ("th"/"en", default "th") picks which fixed
    phrasing to use - see LABEL_LANGS above."""
    usage_mode = data.get("usage_mode", "oral")
    lang = data.get("lang", "th")
    times = _tr_times(data.get("times"), lang)
    if lang == "en":
        if usage_mode == "topical":
            dose_text = f"Apply thinly, {data.get('per_day') or '__'} time(s)/day"
            line2 = "  ".join(times)
            if data.get("every_hr"):
                line2 += f"   every {data['every_hr']} hr"
        elif usage_mode == "drops":
            dose_text = f"{data.get('qty') or '__'} drop(s) per dose"
            if data.get("per_day"):
                dose_text += f"   {data['per_day']} time(s)/day"
            line2 = "  ".join(times)
            if data.get("every_hr"):
                line2 += f"   every {data['every_hr']} hr"
        else:
            unit = _tr_unit(data.get("unit"), lang)
            dose_text = f"Take {data.get('qty') or '__'} {unit} per dose"
            if data.get("per_day"):
                dose_text += f"   {data['per_day']} time(s)/day"
            line2 = _tr_meal(data.get("meal"), lang)
            if times:
                line2 += ("    " if line2 else "") + "  ".join(times)
            if data.get("every_hr"):
                line2 += f"   every {data['every_hr']} hr"
        return dose_text, line2

    if lang == "mm":
        if usage_mode == "topical":
            dose_text = f"တစ်နေ့ {data.get('per_day') or '__'} ကြိမ် ပါးပါးလိမ်းပေးပါ"
            line2 = "  ".join(times)
            if data.get("every_hr"):
                line2 += f"   နာရီ {data['every_hr']} တိုင်း"
        elif usage_mode == "drops":
            dose_text = f"တစ်ခါခတ်လျှင် {data.get('qty') or '__'} စက်"
            if data.get("per_day"):
                dose_text += f"၊ တစ်နေ့ {data['per_day']} ကြိမ် ခတ်ပါ"
            line2 = "  ".join(times)
            if data.get("every_hr"):
                line2 += f"   နာရီ {data['every_hr']} တိုင်း"
        else:
            unit = _tr_unit(data.get("unit"), lang)
            dose_text = f"တစ်ခါသောက်လျှင် {unit} {data.get('qty') or '__'}"
            if data.get("per_day"):
                dose_text += f"၊ တစ်နေ့ {data['per_day']} ကြိမ် သောက်ပါ"
            line2 = _tr_meal(data.get("meal"), lang)
            if times:
                line2 += ("    " if line2 else "") + "  ".join(times)
            if data.get("every_hr"):
                line2 += f"   နာရီ {data['every_hr']} တိုင်း"
        return dose_text, line2

    if usage_mode == "topical":
        # ยาทา: no unit/meal - "ทาบางๆ วันละ N ครั้ง" then times/every_hr only
        dose_text = f"ทาบางๆ วันละ {data.get('per_day') or '__'} ครั้ง"
        line2 = "  ".join(times)
        if data.get("every_hr"):
            line2 += f"   ทุก {data['every_hr']} ชม."
    elif usage_mode == "drops":
        # ยาหยอด: unit fixed to หยด, no meal - "หยอดครั้งละ N หยด วันละ M ครั้ง"
        dose_text = f"หยอดครั้งละ {data.get('qty') or '__'} หยด"
        if data.get("per_day"):
            dose_text += f"   วันละ {data['per_day']} ครั้ง"
        line2 = "  ".join(times)
        if data.get("every_hr"):
            line2 += f"   ทุก {data['every_hr']} ชม."
    else:
        unit = data.get("unit") or "เม็ด"
        dose_text = f"ทานครั้งละ {data.get('qty') or '__'} {unit}"
        if data.get("per_day"):
            dose_text += f"   วันละ {data['per_day']} ครั้ง"
        line2 = data.get("meal") or ""
        if times:
            line2 += ("    " if line2 else "") + "  ".join(times)
        if data.get("every_hr"):
            line2 += f"   ทุก {data['every_hr']} ชม."
    return dose_text, line2


def split_allergy_items(note):
    """Split a free-text allergy note into individual drug-name items, best
    effort - pharmacists type this field freeform with no fixed format, so
    treat common separators (newline, comma, semicolon, Thai full stop) as
    item boundaries. Used to decide whether a patient has exactly one known
    drug allergy (safe to print directly on the label) or several (not
    enough label space - fall back to "ดูแฟ้ม")."""
    if not note:
        return []
    parts = re.split(r"[,\n;、]", note)
    return [p.strip() for p in parts if p.strip()]


# ฉลากเสริม (extra_labels) is chosen from a small fixed preset list per
# usage mode (see EXTRA_LABEL_OPTIONS_BY_MODE) plus one free-typed "อื่นๆ"
# slot - so unlike drug1/drug2/note (genuinely free text, left as typed in
# whatever language on an English label - drug/generic names are usually
# already Latin script anyway, and the pharmacist explains indications
# verbally), the preset phrases can just be hardcoded translations, no AI
# needed. A custom free-typed extra label has no entry here and prints
# as-is, same as drug1/drug2/note.
EXTRA_LABEL_EN = {
    "ทานยาก่อนอาหาร 1/2-1 ชม": "Take 30-60 min before meals",
    "ทานยาหลังอาหารทันที": "Take right after meals",
    "ทานติดต่อกันจนหมด": "Complete the full course",
    "ดื่มน้ำตามมากๆ": "Drink plenty of water",
    "ยานี้อาจทำให้ง่วงซึม": "May cause drowsiness",
    "ห้ามรับประทานพร้อมนม ยาลดกรด": "Avoid with milk/antacids",
    "ทานเมื่อมีอาการ": "Take as needed",
    "หายแล้ว ทาต่ออีก 1 สัปดาห์": "Continue 1 week after healing",
    "ทาต่อเนื่อง 2 สัปดาห์": "Apply continuously for 2 weeks",
    "หยดเมื่อมีอาการ": "Use as needed",
    "หยดต่อเนื่อง 2 สัปดาห์": "Use continuously for 2 weeks",
    "ยาหยอดตาใช้ได้ 3 เดือน": "Discard 3 months after opening",
}
EXTRA_LABEL_MM = {
    "ทานยาก่อนอาหาร 1/2-1 ชม": "ဆေးကို အစားမစားမီ မိနစ် ၃၀ မှ ၁ နာရီကြိုသောက်ပါ",
    "ทานยาหลังอาหารทันที": "အစားစားပြီးချင်း ဆေးချက်ချင်းသောက်ပါ",
    "ทานติดต่อกันจนหมด": "ဆေးကုန်သည်အထိ ဆက်တိုက်သောက်ပါ",
    "ดื่มน้ำตามมากๆ": "ရေများများသောက်ပေးပါ",
    "ยานี้อาจทำให้ง่วงซึม": "ဤဆေးသည် အိပ်ငိုက်စေနိုင်သည်",
    "ห้ามรับประทานพร้อมนม ยาลดกรด": "နို့ သို့မဟုတ် အစာအိမ်ဆေးများနှင့် တွဲမသောက်ရ",
    "ทานเมื่อมีอาการ": "ရောဂါလက္ခဏာရှိမှ သောက်ပါ",
    "หายแล้ว ทาต่ออีก 1 สัปดาห์": "သက်သာပြေကင်းသွားသော်လည်း ၁ ပတ်ဆက်လိမ်းပါ",
    "ทาต่อเนื่อง 2 สัปดาห์": "၂ ပတ်ဆက်တိုက် လိမ်းပေးပါ",
    "หยดเมื่อมีอาการ": "ရောဂါလက္ခဏာရှိမှ ခတ်ပါ",
    "หยดต่อเนื่อง 2 สัปดาห์": "၂ ပတ်ဆက်တိုက် ခတ်ပေးပါ",
    "ยาหยอดตาใช้ได้ 3 เดือน": "မျက်စဉ်းဆေးကို ဖွင့်ပြီး ၃ လအထိသာ သုံးစွဲနိုင်သည်",
}
EXTRA_LABEL_TR = {"en": EXTRA_LABEL_EN, "mm": EXTRA_LABEL_MM}


def _tr_extra_labels(extra_labels, lang):
    table = EXTRA_LABEL_TR.get(lang)
    if not table:
        return list(extra_labels or [])
    return [table.get(e, e) for e in (extra_labels or [])]


EN_STRINGS = {
    "date_label": "Date",
    "patient_label": "Patient",
    "qty_label": "Qty",
    "drug_label": "Drug",
    "generic_label": "Generic Name",
    "warning_text": "Inform Pharmacist: drug allergy, pregnancy, breastfeeding, med condition",
    "align_marker": "breastfeeding,",
    "no_allergy_text": "No known allergy",
    "see_file_text": "Allergy: see file",
    "allergy_prefix": "Allergy: ",
    "pharm_label": "Pharmacist",
}
# Burmese - reviewed/translated by a Burmese-speaking pharmacist contact,
# not machine-translated (same caution as EN_STRINGS).
MM_STRINGS = {
    "date_label": "ရက်စွဲ",
    "patient_label": "လူနာအမည်",
    "qty_label": "အရေအတွက်",
    "drug_label": "ဆေးအမည်",
    "generic_label": "ဆေးအမျိုးအမည်",
    "warning_text": "ဆေးမတည့်၊ရောဂါအခံရှိ၊ကိုယ်ဝန်၊နို့တိုက်မိခင်ဖြစ်ကဆေးမှုအားပြေပါ။",
    "align_marker": "နို့တိုက်",
    "no_allergy_text": "ဆေးဓာတ်မတည့်ခြင်းမရှိပါ",
    "see_file_text": "ဆေးဓာတ်မတည့်မှု - မှတ်တမ်းတွင်ကြည့်ပါ",
    "allergy_prefix": "ဆေးဓာတ်မတည့်မှု - ",
    "pharm_label": "ဆေးဝါးပညာရှင်",
}
LANG_STRINGS = {"en": EN_STRINGS, "mm": MM_STRINGS}


def build_label_image(data, settings):
    lang = data.get("lang", "th")
    L = LANG_STRINGS.get(lang, {})

    def ls(key, thai_default):
        if key in L:
            return L[key]
        if lang == "mm":
            # mmrtext.ttf has no Thai glyphs - if a Burmese string is still
            # missing for this key, fall back to English (Latin renders
            # fine in mmrtext) rather than Thai (would draw as tofu boxes).
            return EN_STRINGS.get(key, thai_default)
        return thai_default

    label_w_px = int(settings["label_w_mm"]) * DOTS_PER_MM
    label_h_px = int(settings["label_h_mm"]) * DOTS_PER_MM

    img = Image.new("L", (label_w_px, label_h_px), 255)
    draw = ImageDraw.Draw(img)

    f_small = find_font(15, lang=lang)
    f_medium = find_font(18, lang=lang)
    f_normal = find_font(19, lang=lang)
    f_label_big = find_font(23, lang=lang)
    # Company name is fixed shop-identity text (from settings, not
    # per-label translated text) and is typically Thai script - always
    # draw it with the Thai/Latin-capable font, even on a Burmese label,
    # since mmrtext.ttf has no Thai glyphs.
    f_bold = find_font(25, bold=True)

    x = 24
    y = 20

    def dotted_field(label, xx, yy, reserved_w):
        """Label followed by a blank line for handwriting (quantity)."""
        draw_thai_text(draw, (xx, yy), label, f_label_big, fill=0)
        lw = draw.textlength(label, font=f_label_big)
        line_x = xx + lw + 6
        draw.line([(line_x, yy + 24), (xx + reserved_w, yy + 24)], fill=0, width=1)

    company_name = settings.get("company_name") or ""
    phone = settings.get("phone") or ""
    draw_thai_text(draw, (x, y), company_name, f_bold, fill=0)

    DATE_RESERVED_W = 180
    if lang in ("en", "mm"):
        today_str = datetime.now().strftime("%d/%m/%Y")
        date_label = ls("date_label", "วันที่")
    else:
        today_str = datetime.now().strftime("%d/%m/") + str(datetime.now().year + 543)
        date_label = "วันที่"
    date_x = label_w_px - x - DATE_RESERVED_W
    draw_thai_text(draw, (date_x, y + 4), date_label, f_label_big, fill=0)
    date_lw = draw.textlength(date_label, font=f_label_big)
    draw_thai_text(draw, (date_x + date_lw + 8, y + 4), today_str, f_label_big, fill=0)

    y += 30
    address_text = f"{settings.get('address_line1') or ''} {settings.get('address_line2') or ''}".strip()
    if address_text:
        # Also fixed Thai shop-identity text - same reasoning as f_bold above.
        address_font = fit_font(draw, address_text, label_w_px - 2 * x, 15, bold=False, min_size=12)
        draw_thai_text(draw, (x, y), address_text, address_font, fill=0)
        if phone:
            addr_w = draw.textlength(address_text, font=address_font)
            draw_thai_text(draw, (x + addr_w + 8, y - 2), f"({phone})", f_medium, fill=0)
        y += 20

    y += 10
    draw.line([(x, y), (label_w_px - x, y)], fill=0, width=2)
    y += 10

    def field(label, value, yy, label_font=f_normal, start_size=30, right_margin=0):
        # value is free-typed (patient name / drug name) - pick its font by
        # inspecting the actual characters, not the label-language dropdown.
        draw_thai_text(draw, (x, yy), label, label_font, fill=0)
        lw = draw.textlength(label, font=label_font)
        value_x = x + lw + 10
        max_w = label_w_px - x - right_margin - value_x
        value_font = fit_font(draw, value or "", max_w, start_size, bold=True, lang=detect_font_lang(value))
        draw_thai_text(draw, (value_x, yy - 3), value or "", value_font, fill=0)

    QTY_RESERVED_W = 150
    field(ls("patient_label", "ชื่อผู้ป่วย"), data["patient_name"], y, right_margin=QTY_RESERVED_W + 10)
    qty_x = label_w_px - x - QTY_RESERVED_W
    label_qty = (data.get("label_qty") or "").strip()
    if not label_qty.isdigit():
        label_qty = ""
    if label_qty:
        # A quantity was actually keyed in for this label - print it
        # directly instead of leaving a blank line for the pharmacist to
        # handwrite it in.
        draw_thai_text(draw, (qty_x, y), f"#{label_qty}", f_label_big, fill=0)
    else:
        dotted_field(ls("qty_label", "จำนวน"), qty_x, y, QTY_RESERVED_W)
    y += 34
    field(ls("drug_label", "ชื่อยา"), data["drug1"], y)
    y += 32
    field(ls("generic_label", "ชื่อยาสามัญ"), data["drug2"], y)
    y += 36

    if data.get("note"):
        # note is free-typed (indication/instructions) - same auto-detect
        # reasoning as patient_name/drug1/drug2 in field() above.
        note_font = fit_font(draw, data["note"], label_w_px - 2 * x, 22, bold=False, min_size=16, lang=detect_font_lang(data["note"]))
        draw_thai_text(draw, (x, y), data["note"], note_font, fill=0)
        y += 28

    dose_text, line2 = compute_dose_lines(data)

    dose_font = fit_font(draw, dose_text, label_w_px - 2 * x, 26, bold=True, lang=lang)
    draw_thai_text(draw, (x, y), dose_text, dose_font, fill=0)
    y += 38

    line2_font = fit_font(draw, line2, label_w_px - 2 * x - QTY_RESERVED_W - 10, 24, bold=True, lang=lang)
    draw_thai_text(draw, (x, y), line2, line2_font, fill=0)
    exp_x = label_w_px - x - QTY_RESERVED_W
    exp_date = (data.get("exp_date") or "").strip()
    if exp_date:
        draw_thai_text(draw, (exp_x, y + 4), f"EXP {exp_date}", f_label_big, fill=0)
    else:
        dotted_field("EXP", exp_x, y + 4, QTY_RESERVED_W)
    y += 40

    extra = _tr_extra_labels(data.get("extra_labels"), lang)
    if extra:
        extra_text = " ".join(f"**{e}**" for e in extra)
        extra_font = fit_font(draw, extra_text, label_w_px - 2 * x, 24, bold=True, lang=lang)
        draw_thai_text(draw, (x, y), extra_text, extra_font, fill=0)
        y += 40

    draw.line([(x, y), (label_w_px - x, y)], fill=0, width=1)
    y += 16
    if lang in ("en", "mm"):
        # This line auto-shrinks via fit_font below to always leave room for
        # the allergy-status text flush right on the same line - no fixed
        # length limit needed here.
        warning_text = ls("warning_text", "")
        align_marker = ls("align_marker", "")
    else:
        warning_text = "แพ้ยา มีโรคประจำตัว ตั้งครรภ์ ให้นมบุตร โปรดแจ้งเภสัชกร"
        align_marker = "ให้นมบุตร"

    # allergy status text - decided before laying out warning_text so the
    # warning line's font can be shrunk to always leave it enough room,
    # rather than drawing warning_text at a fixed size first and hoping the
    # status text still fits next to it (that's how a too-long English
    # phrase used to overlap the status text - see git history).
    no_allergy_text = ls("no_allergy_text", "ไม่แพ้ยา")
    see_file_text = ls("see_file_text", "แพ้ยา:ดูแฟ้ม")
    if not data.get("has_allergy"):
        status_text = no_allergy_text
    else:
        allergy_drug_name = (data.get("allergy_drug_name") or "").strip()
        status_text = (
            (f"{ls('allergy_prefix', 'แพ้ยา:')}{allergy_drug_name}")
            if allergy_drug_name else see_file_text
        )

    STATUS_RESERVED_W = 160
    warning_font = fit_font(
        draw, warning_text, label_w_px - 2 * x - STATUS_RESERVED_W, 18, min_size=11, lang=lang,
    )
    draw_thai_text(draw, (x, y), warning_text, warning_font, fill=0)

    status_font = f_medium
    status_w = draw.textlength(status_text, font=status_font)
    if status_w > STATUS_RESERVED_W - 10:
        # Even the generic fallback text doesn't fit the reserved space at
        # this label size - shrink it too rather than let it run off the
        # edge of the sticker.
        status_font = fit_font(draw, status_text, STATUS_RESERVED_W - 10, 18, min_size=11, lang=lang)
        status_w = draw.textlength(status_text, font=status_font)
    status_x = label_w_px - x - status_w
    draw_thai_text(draw, (status_x, y), status_text, status_font, fill=0)
    underline_y = y + 18
    draw.line([(status_x, underline_y), (status_x + status_w, underline_y)], fill=0, width=1)
    y += 22

    # pharmacist name on its own row below the warning line, starting at the
    # same x as align_marker on the line above - needs the extra width for
    # pharmacists who want full name + surname + license number
    # pharm_label is per-language translated text (drawn with the lang-aware
    # font); pharmacist_names is fixed Thai shop-config text (always drawn
    # with the Thai/Latin-capable font, same reasoning as f_bold/
    # address_font above) - two separate draw calls since mmrtext.ttf has
    # no Thai glyphs to render pharmacist_names with on a Burmese label.
    pharmacist_names = settings.get("pharmacist_names") or ""
    pharm_label = ls("pharm_label", "เภสัชกร")
    if pharmacist_names:
        pharm_label_text = f"{pharm_label}: "
        pharm_font = find_font(18, lang=lang)  # 15 * 1.2
        pharm_name_font = find_font(18)  # always Thai/Latin-capable
        prefix_w = draw.textlength(warning_text.split(align_marker)[0], font=warning_font)
        draw_thai_text(draw, (x + prefix_w, y), pharm_label_text, pharm_font, fill=0)
        label_w = draw.textlength(pharm_label_text, font=pharm_font)
        draw_thai_text(draw, (x + prefix_w + label_w, y), pharmacist_names, pharm_name_font, fill=0)

    return img


# A4 sheet mode - some shops don't have a thermal label printer, only a
# regular office printer, and want to print several labels per A4 sheet
# and cut them apart by hand (or print to "Microsoft Print to PDF", which
# is just another entry in the printer list - no separate PDF export code
# needed, the existing print_image() pipeline already handles it since
# that "printer" is what generates the PDF file).
A4_W_MM = 210
A4_H_MM = 297
A4_MARGIN_MM = 5
A4_GAP_MM = 3
A4_COLUMNS = 2


A4_DASH_LEN_PX = 10  # ~1.2mm at 8 dots/mm - short dash + gap reads as a perforated cut line
A4_DASH_GAP_PX = 6


def _draw_dashed_rect(draw, box, dash_len=A4_DASH_LEN_PX, gap_len=A4_DASH_GAP_PX, fill=0, width=1):
    """Perforation-style cut guide (dashes, not a solid line) around one
    label's edge on the A4 sheet - drawn on the page, not on the label
    image itself, so it never overlaps the label's own content."""
    x0, y0, x1, y1 = box
    step = dash_len + gap_len
    xx = x0
    while xx < x1:
        seg_end = min(xx + dash_len, x1)
        draw.line([(xx, y0), (seg_end, y0)], fill=fill, width=width)
        draw.line([(xx, y1), (seg_end, y1)], fill=fill, width=width)
        xx += step
    yy = y0
    while yy < y1:
        seg_end = min(yy + dash_len, y1)
        draw.line([(x0, yy), (x0, seg_end)], fill=fill, width=width)
        draw.line([(x1, yy), (x1, seg_end)], fill=fill, width=width)
        yy += step


def build_a4_pages(label_images):
    """Tiles same-size label images (already rendered at their configured
    physical label_w_mm x label_h_mm) into a grid of up to A4_COLUMNS
    columns per A4 page, as many rows as fit - so each label prints at its
    real physical size on the sheet, not stretched to fill the page. Falls
    back to 1 column if the configured label is too wide for 2 side by
    side. Each label gets a dashed "รอยปรุ" cut guide around its edge so
    it's obvious where to tear/cut apart by hand. Returns a list of full
    A4-page PIL images, one per page needed."""
    if not label_images:
        return []
    label_w_px, label_h_px = label_images[0].size
    margin_px = A4_MARGIN_MM * DOTS_PER_MM
    gap_px = A4_GAP_MM * DOTS_PER_MM
    page_w_px = A4_W_MM * DOTS_PER_MM
    page_h_px = A4_H_MM * DOTS_PER_MM

    usable_w = page_w_px - 2 * margin_px
    usable_h = page_h_px - 2 * margin_px
    columns = max(1, min(A4_COLUMNS, (usable_w + gap_px) // (label_w_px + gap_px)))
    rows_per_page = max(1, (usable_h + gap_px) // (label_h_px + gap_px))
    per_page = int(columns * rows_per_page)

    pages = []
    for start in range(0, len(label_images), per_page):
        chunk = label_images[start:start + per_page]
        page = Image.new("L", (page_w_px, page_h_px), 255)
        page_draw = ImageDraw.Draw(page)
        for i, lbl in enumerate(chunk):
            col = i % columns
            row = i // columns
            px = int(margin_px + col * (label_w_px + gap_px))
            py = int(margin_px + row * (label_h_px + gap_px))
            page.paste(lbl, (px, py))
            _draw_dashed_rect(page_draw, (px, py, px + label_w_px, py + label_h_px))
        pages.append(page)
    return pages


LANG_NAMES_FOR_GROK_PROMPT = {"en": "English", "mm": "Burmese (Myanmar script)"}


def translate_note_via_grok(note_text, target_lang):
    """Translate a free-typed Indication/note field via Grok (xAI), reusing
    the same provider client and stored API key as the existing "AI ช่วยค้น
    ข้อมูล" feature (ai_assist.call_xai + app_settings' xai_api_key - see
    ai_assist.py / open_ai_assist_dialog). Returns (translated_text_or_None,
    error_message_or_None) - never raises, since a failed/unconfigured
    translation must fall back to the original Thai text rather than block
    preview/printing."""
    settings = app_settings.load_settings()
    api_key = (settings.get("xai_api_key") or "").strip()
    if not api_key:
        return None, "ยังไม่ได้ตั้งค่า Grok (xAI) API key - ตั้งค่าได้ที่ปุ่ม 🤖 AI ช่วยค้นข้อมูล ก่อน"
    lang_name = LANG_NAMES_FOR_GROK_PROMPT.get(target_lang, target_lang)
    prompt = (
        f"Translate the following pharmacy label indication/instruction text from Thai to {lang_name}. "
        f"Reply with ONLY the translated text - no explanation, no quotes, no extra commentary:\n\n{note_text}"
    )
    ok, text = ai_assist.call_xai(api_key, prompt)
    if not ok:
        return None, text
    return text.strip(), None


def print_image(img, printer_name=None):
    import win32print
    import win32ui

    if not printer_name:
        printer_name = app_settings.load_settings().get("printer_name") or ""
    if not printer_name:
        raise RuntimeError("ยังไม่ได้ตั้งค่าเครื่องพิมพ์ - กรุณาไปที่เมนู ⚙️ ตั้งค่า ก่อน")

    hdc = win32ui.CreateDC()
    hdc.CreatePrinterDC(printer_name)
    hdc.StartDoc("ฉลากยา")
    hdc.StartPage()

    dib = ImageWin.Dib(img)
    printable_w = hdc.GetDeviceCaps(8)   # HORZRES
    printable_h = hdc.GetDeviceCaps(10)  # VERTRES
    dib.draw(hdc.GetHandleOutput(), (0, 0, printable_w, printable_h))

    hdc.EndPage()
    hdc.EndDoc()
    hdc.DeleteDC()


def build_settings_dialog(parent, first_run=False):
    """Settings screen: printer, label size, company/pharmacist info. Shown
    automatically on first run (no settings.json yet), reopenable any time
    via the ⚙️ ตั้งค่า button."""
    settings = app_settings.load_settings()
    dialog_win = tk.Toplevel(parent)
    dialog_win.title("ตั้งค่าเริ่มต้น (ครั้งแรก)" if first_run else "ตั้งค่า")
    dialog_win.geometry(f"{fs(440)}x{fs(600)}")
    dialog_win.transient(parent)
    dialog_win.grab_set()

    pad = {"padx": fs(10), "pady": fs(4)}

    # Footer (save + admin buttons) packed FIRST with side="bottom" so it
    # claims its space at the bottom of the window before the scrollable
    # area below gets whatever's left - this is what keeps "บันทึก" always
    # visible no matter how many fields (e.g. the 3 API key entries) get
    # added above it, instead of the footer silently getting pushed off the
    # bottom of a fixed-height window.
    footer = tk.Frame(dialog_win)
    footer.pack(side="bottom", fill="x")

    scroll_container = tk.Frame(dialog_win)
    scroll_container.pack(side="top", fill="both", expand=True)
    canvas = tk.Canvas(scroll_container, highlightthickness=0)
    scrollbar = tk.Scrollbar(scroll_container, orient="vertical", command=canvas.yview)
    win = tk.Frame(canvas)  # `win` = the scrollable content frame, all fields below attach here
    win.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
    canvas_window = canvas.create_window((0, 0), window=win, anchor="nw")
    canvas.bind("<Configure>", lambda e: canvas.itemconfig(canvas_window, width=e.width))
    canvas.configure(yscrollcommand=scrollbar.set)
    canvas.pack(side="left", fill="both", expand=True)
    scrollbar.pack(side="right", fill="y")

    def _on_mousewheel(event):
        canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")

    canvas.bind("<Enter>", lambda e: canvas.bind_all("<MouseWheel>", _on_mousewheel))
    canvas.bind("<Leave>", lambda e: canvas.unbind_all("<MouseWheel>"))

    if first_run:
        tk.Label(
            win, text="ตั้งค่าครั้งแรก - กรอกข้อมูลร้านและเลือกเครื่องพิมพ์ก่อนเริ่มใช้งาน",
            font=("Tahoma", fs(10), "bold"), fg="#a00", wraplength=fs(400), justify="left",
        ).pack(anchor="w", **pad)

    tk.Label(win, text="เครื่องพิมพ์สติ๊กเกอร์ *", font=("Tahoma", fs(10), "bold")).pack(anchor="w", **pad)
    printers = app_settings.list_printers()
    printer_var = tk.StringVar(value=settings["printer_name"] or (printers[0] if printers else ""))
    printer_menu_values = printers or ["(ไม่พบเครื่องพิมพ์ที่ติดตั้งไว้)"]
    tk.OptionMenu(win, printer_var, *printer_menu_values).pack(fill="x", **pad)

    size_frame = tk.Frame(win)
    size_frame.pack(fill="x", **pad)
    tk.Label(size_frame, text="ขนาดฉลาก (มม.) กว้าง", font=("Tahoma", fs(10))).grid(row=0, column=0)
    w_var = tk.StringVar(value=str(settings["label_w_mm"]))
    tk.Entry(size_frame, textvariable=w_var, width=5, font=("Tahoma", fs(10))).grid(row=0, column=1, padx=fs(4))
    tk.Label(size_frame, text="สูง", font=("Tahoma", fs(10))).grid(row=0, column=2)
    h_var = tk.StringVar(value=str(settings["label_h_mm"]))
    tk.Entry(size_frame, textvariable=h_var, width=5, font=("Tahoma", fs(10))).grid(row=0, column=3, padx=fs(4))

    tk.Label(win, text="ประเภทกระดาษ", font=("Tahoma", fs(10), "bold")).pack(anchor="w", **pad)
    PAPER_MODE_LABELS = {"thermal": "ฉลากม้วน (Thermal)", "a4": "A4 (2 คอลัมน์ต่อแผ่น, ตัดแยกเอง)"}
    PAPER_MODE_KEYS_BY_LABEL = {v: k for k, v in PAPER_MODE_LABELS.items()}
    paper_mode_var = tk.StringVar(value=PAPER_MODE_LABELS[settings.get("paper_mode", "thermal")])
    paper_mode_combo = ttk.Combobox(
        win, textvariable=paper_mode_var, values=list(PAPER_MODE_LABELS.values()),
        state="readonly", font=("Tahoma", fs(10)),
    )
    paper_mode_combo.pack(fill="x", **pad)
    tk.Label(
        win, text="A4: เอาไว้เผื่อร้านที่ไม่มีเครื่องพิมพ์สติ๊กเกอร์ - พิมพ์ฉลากขนาดจริง 2 ใบต่อแถวลงกระดาษ "
                   "A4 แล้วตัดแยกเอง หรือเลือกเครื่องพิมพ์เป็น \"Microsoft Print to PDF\" ด้านบนเพื่อ export เป็น PDF แทนได้เลย",
        font=("Tahoma", fs(8)), fg="#666", wraplength=fs(400), justify="left",
    ).pack(anchor="w", padx=fs(10))

    tk.Label(win, text="ชื่อร้าน/บริษัท", font=("Tahoma", fs(10), "bold")).pack(anchor="w", **pad)
    name_var = tk.StringVar(value=settings["company_name"])
    tk.Entry(win, textvariable=name_var, font=("Tahoma", fs(11))).pack(fill="x", **pad)

    tk.Label(win, text="ที่อยู่ บรรทัด 1", font=("Tahoma", fs(10), "bold")).pack(anchor="w", **pad)
    addr1_var = tk.StringVar(value=settings["address_line1"])
    tk.Entry(win, textvariable=addr1_var, font=("Tahoma", fs(11))).pack(fill="x", **pad)

    tk.Label(win, text="ที่อยู่ บรรทัด 2", font=("Tahoma", fs(10), "bold")).pack(anchor="w", **pad)
    addr2_var = tk.StringVar(value=settings["address_line2"])
    tk.Entry(win, textvariable=addr2_var, font=("Tahoma", fs(11))).pack(fill="x", **pad)

    tk.Label(win, text="เบอร์โทร", font=("Tahoma", fs(10), "bold")).pack(anchor="w", **pad)
    phone_var = tk.StringVar(value=settings["phone"])
    tk.Entry(win, textvariable=phone_var, font=("Tahoma", fs(11))).pack(fill="x", **pad)

    tk.Label(win, text="ชื่อเภสัชกร (พิมพ์ตามที่ต้องการ เช่น สมชาย//สมหญิง)",
             font=("Tahoma", fs(10), "bold"), wraplength=fs(400), justify="left").pack(anchor="w", **pad)
    pharm_var = tk.StringVar(value=settings["pharmacist_names"])
    tk.Entry(win, textvariable=pharm_var, font=("Tahoma", fs(11))).pack(fill="x", **pad)

    tk.Label(
        win, text="API Key สำหรับ AI ช่วยค้นข้อมูล (ไม่บังคับ - ใส่เฉพาะตัวที่จะใช้)",
        font=("Tahoma", fs(10), "bold"), wraplength=fs(400), justify="left",
    ).pack(anchor="w", **pad)
    ai_key_vars = {}
    for provider_key, provider_info in ai_assist.PROVIDERS.items():
        tk.Label(win, text=provider_info["label"], font=("Tahoma", fs(9))).pack(anchor="w", padx=fs(10))
        v = tk.StringVar(value=settings.get(provider_info["key_field"], ""))
        ai_key_vars[provider_key] = v
        tk.Entry(win, textvariable=v, font=("Tahoma", fs(10)), show="•").pack(fill="x", padx=fs(10), pady=(0, fs(4)))

    status_var = tk.StringVar(value="")
    tk.Label(footer, textvariable=status_var, font=("Tahoma", fs(9)), fg="#a00",
             wraplength=fs(400), justify="left").pack(padx=fs(10))

    def on_save():
        if not printer_var.get().strip():
            status_var.set("กรุณาเลือกเครื่องพิมพ์")
            return
        try:
            w = int(w_var.get().strip())
            h = int(h_var.get().strip())
        except ValueError:
            status_var.set("ขนาดฉลากต้องเป็นตัวเลข")
            return
        app_settings.save_settings({
            "printer_name": printer_var.get().strip(),
            "label_w_mm": w, "label_h_mm": h,
            "paper_mode": PAPER_MODE_KEYS_BY_LABEL.get(paper_mode_var.get(), "thermal"),
            "company_name": name_var.get().strip(),
            "address_line1": addr1_var.get().strip(),
            "address_line2": addr2_var.get().strip(),
            "phone": phone_var.get().strip(),
            "pharmacist_names": pharm_var.get().strip(),
            **{ai_assist.PROVIDERS[k]["key_field"]: v.get().strip() for k, v in ai_key_vars.items()},
        })
        dialog_win.destroy()

    tk.Button(
        footer, text="💾 บันทึก", font=("Tahoma", fs(11), "bold"), bg="#1a7a4a", fg="white", command=on_save,
    ).pack(pady=fs(12))

    if not first_run:
        def on_clear_db():
            count = storage.count_templates()
            if count == 0:
                messagebox.showinfo("แจ้งเตือน", "ยังไม่มีข้อมูลยาในเครื่องนี้เลย", parent=dialog_win)
                return
            if not messagebox.askyesno(
                "ยืนยันการล้างข้อมูล",
                f"จะลบข้อมูลยาทั้งหมด {count} รายการออกจากเครื่องนี้ถาวร\n"
                "(รายการยาที่กรอกวิธีใช้ไว้ทั้งหมดจะหายไป กู้คืนไม่ได้)\n\n"
                "ยืนยันที่จะลบหรือไม่?",
                icon="warning", parent=dialog_win,
            ):
                return
            if not messagebox.askyesno("ยืนยันอีกครั้ง", "แน่ใจจริงๆ ใช่ไหม? ข้อมูลจะหายทั้งหมดและกู้คืนไม่ได้",
                                        icon="warning", parent=dialog_win):
                return
            removed = storage.clear_all_templates()
            messagebox.showinfo("สำเร็จ", f"ลบข้อมูลยาแล้ว {removed} รายการ", parent=dialog_win)

        def on_show_all_drugs():
            names = storage.list_all_template_names()
            list_win = tk.Toplevel(dialog_win)
            list_win.title(f"ยาทั้งหมดในเครื่องนี้ ({len(names)} รายการ)")
            list_win.geometry(f"{fs(380)}x{fs(480)}")
            list_win.transient(dialog_win)
            list_win.grab_set()

            tk.Label(
                list_win, text=f"ยาทั้งหมดในเครื่องนี้ ({len(names)} รายการ)",
                font=("Tahoma", fs(11), "bold"),
            ).pack(anchor="w", padx=fs(10), pady=(fs(10), fs(4)))

            list_frame = tk.Frame(list_win)
            list_frame.pack(fill="both", expand=True, padx=fs(10), pady=(0, fs(8)))
            scrollbar = tk.Scrollbar(list_frame, orient="vertical")
            listbox = tk.Listbox(list_frame, font=("Tahoma", fs(10)), yscrollcommand=scrollbar.set)
            scrollbar.config(command=listbox.yview)
            scrollbar.pack(side="right", fill="y")
            listbox.pack(side="left", fill="both", expand=True)
            if names:
                for n in names:
                    listbox.insert(tk.END, n)
            else:
                listbox.insert(tk.END, "(ยังไม่มีข้อมูลยาในเครื่องนี้เลย)")

            tk.Button(list_win, text="ปิด", font=("Tahoma", fs(10)), command=list_win.destroy).pack(
                pady=(0, fs(10)))
            list_win.lift()
            list_win.focus_force()

        clear_row = tk.Frame(footer)
        clear_row.pack(pady=(0, fs(10)))
        tk.Button(
            clear_row, text="📋 แสดงยาทั้งหมด", font=("Tahoma", fs(9)),
            command=on_show_all_drugs,
        ).pack(side="left", padx=(0, fs(6)))
        tk.Button(
            clear_row, text="🗑 ล้าง DB (ลบข้อมูลยาทั้งหมด)", font=("Tahoma", fs(9)),
            fg="#a02020", command=on_clear_db,
        ).pack(side="left")

        def on_export():
            path = filedialog.asksaveasfilename(
                title="บันทึกไฟล์สำรองข้อมูล", defaultextension=".zip",
                filetypes=[("Zip files", "*.zip")],
                initialfile=f"labelprinter_backup_{datetime.now().strftime('%Y%m%d')}.zip",
                parent=dialog_win,
            )
            if not path:
                return
            try:
                export_backup(path)
            except Exception as e:
                messagebox.showerror("ผิดพลาด", f"Export ไม่สำเร็จ: {e}", parent=dialog_win)
                return
            messagebox.showinfo("สำเร็จ", f"บันทึกไฟล์สำรองไว้ที่:\n{path}\n\nเอาไฟล์นี้ไปเปิดที่เครื่องใหม่แล้วกด Import ได้เลย", parent=dialog_win)

        def on_import():
            path = filedialog.askopenfilename(
                title="เลือกไฟล์สำรองข้อมูล (.zip)", filetypes=[("Zip files", "*.zip")], parent=dialog_win,
            )
            if not path:
                return
            if not messagebox.askyesno(
                "ยืนยัน",
                "การ Import จะเขียนทับข้อมูลยาทั้งหมด, Favorite, และข้อมูลร้าน "
                "(ยกเว้นเครื่องพิมพ์) ในเครื่องนี้ทันที\nข้อมูลเดิมจะหายไปถ้าไม่ได้สำรองไว้ก่อน\n\n"
                "ยืนยันที่จะ Import หรือไม่?",
                icon="warning", parent=dialog_win,
            ):
                return
            try:
                import_backup(path)
            except Exception as e:
                messagebox.showerror("ผิดพลาด", f"Import ไม่สำเร็จ: {e}", parent=dialog_win)
                return
            messagebox.showinfo(
                "สำเร็จ", "Import ข้อมูลสำเร็จ\n\nกรุณาปิดโปรแกรมแล้วเปิดใหม่เพื่อให้ข้อมูลอัปเดตครบถ้วน", parent=dialog_win)
            dialog_win.destroy()

        backup_row = tk.Frame(footer)
        backup_row.pack(pady=(0, fs(10)))
        tk.Button(
            backup_row, text="📤 Export ข้อมูลทั้งหมด", font=("Tahoma", fs(9), "bold"),
            bg="#1a5a9a", fg="white", command=on_export,
        ).pack(side="left", padx=fs(4))
        tk.Button(
            backup_row, text="📥 Import จากไฟล์สำรอง", font=("Tahoma", fs(9), "bold"),
            bg="#1a5a9a", fg="white", command=on_import,
        ).pack(side="left", padx=fs(4))

    dialog_win.lift()
    dialog_win.focus_force()
    return dialog_win


class LabelApp:
    def __init__(self, root):
        self.root = root
        root.title(f"พิมพ์ฉลากยา v{APP_VERSION}")
        root.geometry(f"{fs(1050)}x{fs(700)}")  # fallback size if un-maximized later
        root.state("zoomed")  # start maximized - the toolbar row doesn't fit at the old default size
        # ttk.Combobox's dropdown popup is a separate internal Listbox that
        # doesn't inherit the widget's own font= option - has to be set via
        # the option database to make dropdown list text bigger too.
        root.option_add("*TCombobox*Listbox.font", ("Tahoma", fs(12)))

        self.search_results = []
        self.selected_drugs = []  # list of dicts, see add_drug()
        self._search_after_id = None
        self.favorites = load_favorites()  # name -> list of drug dicts
        self._queue_patient_name = None
        self._queue_patient_phone = None
        self._queue_patient_id = None
        self.NO_CUSTOMER_TEXT = "รอชื่อลูกค้า (optional)"

        # Local-LAN server so staff phones on the same WiFi can submit drugs
        # into the print queue - no cloud dependency, matches the "one PC,
        # one printer" assumption of this standalone build. Best-effort: if
        # every port in range is somehow taken, queue_url stays None and the
        # queue button just explains that instead of crashing the app.
        queue_ip, queue_port = local_server.start_server()
        self.queue_url = f"http://{queue_ip}:{queue_port}" if queue_ip else None

        pad = {"padx": fs(10), "pady": fs(4)}

        main_frame = tk.Frame(root)
        main_frame.pack(fill="both", expand=True)

        left_frame = tk.Frame(main_frame)
        left_frame.pack(side="left", fill="both", expand=True)

        right_frame = tk.Frame(main_frame, width=fs(230), bg="#f0f0f0", relief="groove", bd=1)
        right_frame.pack(side="right", fill="y")
        right_frame.pack_propagate(False)

        # ---------------------------------------------------------- left: search + selected list

        search_header = tk.Frame(left_frame)
        search_header.pack(fill="x", **pad)
        tk.Label(search_header, text="ค้นหายาที่บันทึกไว้แล้วดับเบิลคลิกเพื่อเพิ่มเข้ารายการ",
                 font=("Tahoma", fs(10), "bold")).pack(side="left", anchor="w")
        tk.Button(
            search_header, text="⚙️ ตั้งค่า", font=("Tahoma", fs(9)),
            command=lambda: build_settings_dialog(self.root),
        ).pack(side="right")
        tk.Button(
            search_header, text="📥 Import จาก Excel", font=("Tahoma", fs(9)),
            command=self.open_import_excel_dialog,
        ).pack(side="right", padx=(0, fs(6)))
        tk.Button(
            search_header, text="📱 คิวจากมือถือ", font=("Tahoma", fs(9), "bold"),
            bg="#5a5a9a", fg="white", command=self.open_queue_dialog,
        ).pack(side="right", padx=(0, fs(6)))
        tk.Button(
            search_header, text="📜 แฟ้มประวัติการจ่ายยา", font=("Tahoma", fs(9)),
            command=self.open_print_history_dialog,
        ).pack(side="right", padx=(0, fs(6)))
        tk.Button(
            search_header, text="🗂 ประวัติผู้ป่วย", font=("Tahoma", fs(9)),
            command=self.open_patient_profile_dialog,
        ).pack(side="right", padx=(0, fs(6)))
        tk.Button(
            search_header, text="🤖 AI ช่วยค้นข้อมูล", font=("Tahoma", fs(9)),
            command=self.open_ai_assist_dialog,
        ).pack(side="right", padx=(0, fs(6)))

        search_row = tk.Frame(left_frame)
        search_row.pack(fill="x", **pad)
        self.search_var = tk.StringVar()
        self.search_var.trace_add("write", self.on_search_change)
        search_entry = tk.Entry(search_row, textvariable=self.search_var, font=("Tahoma", fs(11)))
        search_entry.pack(side="left", fill="x", expand=True)
        search_entry.bind("<Return>", self.on_search_enter)
        tk.Button(
            search_row, text="✕", font=("Tahoma", fs(9), "bold"), fg="white", bg="#555555", width=2,
            command=lambda: self.search_var.set(""),
        ).pack(side="left", padx=(fs(4), 0))
        tk.Button(
            search_row, text="+ เพิ่มยาใหม่", font=("Tahoma", fs(9), "bold"),
            bg="#1a5a9a", fg="white", command=self.add_new_drug,
        ).pack(side="left", padx=(fs(4), 0))
        results_container = tk.Frame(left_frame, height=fs(140), bg="white")
        results_container.pack(fill="x", **pad)
        results_container.pack_propagate(False)
        self._results_canvas = tk.Canvas(results_container, highlightthickness=0, bg="white")
        results_scrollbar = tk.Scrollbar(results_container, orient="vertical", command=self._results_canvas.yview)
        self.results_frame = tk.Frame(self._results_canvas, bg="white")
        self.results_frame.bind(
            "<Configure>",
            lambda e: self._results_canvas.configure(scrollregion=self._results_canvas.bbox("all")),
        )
        self._results_canvas_window = self._results_canvas.create_window((0, 0), window=self.results_frame, anchor="nw")
        self._results_canvas.bind(
            "<Configure>",
            lambda e: self._results_canvas.itemconfig(self._results_canvas_window, width=e.width),
        )
        self._results_canvas.configure(yscrollcommand=results_scrollbar.set)
        self._results_canvas.pack(side="left", fill="both", expand=True)
        results_scrollbar.pack(side="right", fill="y")

        list_header = tk.Frame(left_frame)
        list_header.pack(fill="x", **pad)
        tk.Label(list_header, text="รายการยาที่จะพิมพ์ (เขียว = บันทึกไว้แล้ว, ส้ม = แก้ไขแต่ยังไม่ได้ save, "
                                    "เหลือง = ใช้ค่าเฉพาะ Favorite นี้, แดง = ยังไม่มีข้อมูล, ดับเบิลคลิกเพื่อแก้ไข)",
                 font=("Tahoma", fs(9), "bold"), wraplength=fs(340), justify="left").pack(side="left", anchor="w")
        tk.Button(
            list_header, text="🗑 ล้างทั้งหมด", font=("Tahoma", fs(9)),
            command=self.clear_all_drugs,
        ).pack(side="right", anchor="n", padx=(fs(6), 0))
        self.allergy_var = tk.BooleanVar(value=False)
        allergy_frame = tk.Frame(list_header)
        allergy_frame.pack(side="right", anchor="n", padx=(fs(6), 0))
        tk.Checkbutton(
            allergy_frame, variable=self.allergy_var, font=("Tahoma", fs(9), "bold"), fg="#b03a2e",
        ).pack(side="left")
        tk.Button(
            allergy_frame, text="⚠ แพ้ยา", font=("Tahoma", fs(9), "bold"), fg="#b03a2e",
            relief="flat", bd=0, cursor="hand2",
            command=lambda: self.open_patient_profile_dialog(preload_patient_id=self._queue_patient_id),
        ).pack(side="left")

        # Pick a returning customer straight from the existing "แฟ้มประวัติ
        # การจ่ายยา" dialog (it already has full search/browsing) instead of
        # building a separate lookup UI - just adds a "เลือกชื่อนี้" action
        # in there that carries the name+phone back here. The picked-name
        # display sits directly under this button so it's obviously tied to it.
        pick_customer_col = tk.Frame(list_header)
        pick_customer_col.pack(side="right", anchor="n", padx=(fs(6), 0))
        tk.Button(
            pick_customer_col, text="👤 เลือกชื่อลูกค้า", font=("Tahoma", fs(9)),
            command=lambda: self.open_print_history_dialog(pick_mode=True),
        ).pack(side="top")
        customer_row = tk.Frame(pick_customer_col)
        customer_row.pack(side="top", fill="x")
        self.selected_customer_var = tk.StringVar(value=self.NO_CUSTOMER_TEXT)
        tk.Label(
            customer_row, textvariable=self.selected_customer_var, font=("Tahoma", fs(10), "bold"),
            fg="#0a3d7a", anchor="w",
        ).pack(side="left")
        tk.Button(
            customer_row, text="✕", font=("Tahoma", fs(7), "bold"), fg="white", bg="#555555", width=2,
            command=self.clear_selected_customer,
        ).pack(side="left", padx=(fs(2), 0))

        list_container = tk.Frame(left_frame)
        list_container.pack(fill="both", expand=True, padx=fs(10), pady=fs(4))
        self._list_canvas = tk.Canvas(list_container, highlightthickness=0)
        list_scrollbar = tk.Scrollbar(list_container, orient="vertical", command=self._list_canvas.yview)
        self.list_frame = tk.Frame(self._list_canvas)
        self.list_frame.bind(
            "<Configure>",
            lambda e: self._list_canvas.configure(scrollregion=self._list_canvas.bbox("all")),
        )
        self._list_canvas_window = self._list_canvas.create_window((0, 0), window=self.list_frame, anchor="nw")
        self._list_canvas.bind(
            "<Configure>",
            lambda e: self._list_canvas.itemconfig(self._list_canvas_window, width=e.width),
        )
        self._list_canvas.configure(yscrollcommand=list_scrollbar.set)
        self._list_canvas.pack(side="left", fill="both", expand=True)
        list_scrollbar.pack(side="right", fill="y")

        self.confirm_btn = tk.Button(
            left_frame, text="ยืนยันรายการยา → ใส่ชื่อผู้ป่วย", font=("Tahoma", fs(12), "bold"),
            bg="#1a7a4a", fg="white", command=self.on_confirm,
        )
        self.confirm_btn.pack(pady=(fs(10), fs(12)))

        self.status_var = tk.StringVar(value="")
        tk.Label(left_frame, textvariable=self.status_var, font=("Tahoma", fs(9)), fg="#070",
                 wraplength=fs(420), justify="left").pack(anchor="w", padx=fs(10), pady=(0, fs(8)))

        # ---------------------------------------------------------- right: favorites panel

        tk.Label(right_frame, text="⭐ Favorite ชุดยา", font=("Tahoma", fs(11), "bold"), bg="#f0f0f0").pack(pady=(fs(10), fs(6)))

        tk.Button(
            right_frame, text="+ บันทึกรายการนี้เป็น Favorite", font=("Tahoma", fs(9), "bold"),
            bg="#1a7a4a", fg="white", wraplength=fs(190), command=self.on_save_favorite,
        ).pack(fill="x", padx=fs(8), pady=(0, fs(8)))

        fav_search_row = tk.Frame(right_frame, bg="#f0f0f0")
        fav_search_row.pack(fill="x", padx=fs(8), pady=(0, fs(6)))
        self.fav_search_var = tk.StringVar()
        self.fav_search_var.trace_add("write", lambda *a: self.refresh_fav_buttons())
        tk.Entry(fav_search_row, textvariable=self.fav_search_var, font=("Tahoma", fs(9))).pack(side="left", fill="x", expand=True)
        tk.Button(
            fav_search_row, text="✕", font=("Tahoma", fs(8), "bold"), fg="white", bg="#555555", width=2,
            command=lambda: self.fav_search_var.set(""),
        ).pack(side="left", padx=(fs(4), 0))

        fav_scroll_container = tk.Frame(right_frame, bg="#f0f0f0")
        fav_scroll_container.pack(fill="both", expand=True, padx=fs(8), pady=(0, fs(8)))
        self._fav_canvas = tk.Canvas(fav_scroll_container, highlightthickness=0, bg="#f0f0f0")
        fav_scrollbar = tk.Scrollbar(fav_scroll_container, orient="vertical", command=self._fav_canvas.yview)
        self.fav_button_frame = tk.Frame(self._fav_canvas, bg="#f0f0f0")
        self.fav_button_frame.bind(
            "<Configure>",
            lambda e: self._fav_canvas.configure(scrollregion=self._fav_canvas.bbox("all")),
        )
        self._fav_canvas_window = self._fav_canvas.create_window((0, 0), window=self.fav_button_frame, anchor="nw")
        self._fav_canvas.bind(
            "<Configure>",
            lambda e: self._fav_canvas.itemconfig(self._fav_canvas_window, width=e.width),
        )
        self._fav_canvas.configure(yscrollcommand=fav_scrollbar.set)
        self._fav_canvas.pack(side="left", fill="both", expand=True)
        fav_scrollbar.pack(side="right", fill="y")

        self.refresh_fav_buttons()

    def _bind_mousewheel(self, widget, canvas):
        """Bind wheel-scroll directly on widget and every descendant, so scrolling
        works no matter which child (button/label/row frame) the cursor is over."""
        widget.bind("<MouseWheel>", lambda e: canvas.yview_scroll(int(-1 * (e.delta / 120)), "units"))
        for child in widget.winfo_children():
            self._bind_mousewheel(child, canvas)

    # ---------------------------------------------------------------- search

    def on_search_change(self, *args):
        if self._search_after_id:
            self.root.after_cancel(self._search_after_id)
        self._search_after_id = self.root.after(300, self.do_search)

    def on_search_enter(self, event=None):
        """A barcode scanner is just a fast keyboard that types the code
        then sends Enter - so Enter in this box is the natural place to
        check "does this exactly match one barcode?" and skip straight to
        adding it, instead of making the pharmacist scan then still have to
        double-click a search result."""
        term = self.search_var.get().strip()
        if not term:
            return
        match = storage.find_template_by_barcode(term)
        if not match:
            return
        self.search_var.set("")

        def worker():
            try:
                info = get_product_med_info(match["idproduct"])
            except Exception as e:
                self.root.after(0, lambda: self.status_var.set(f"เกิดข้อผิดพลาด: {e}"))
                return
            self.root.after(0, lambda: self.add_drug(match, info))

        threading.Thread(target=worker, daemon=True).start()

    def do_search(self):
        term = self.search_var.get().strip()
        self.search_results = []
        self._render_search_results()
        if not term:
            return

        def worker():
            try:
                results = search_products(term)
            except Exception as e:
                self.root.after(0, lambda: self.status_var.set(f"ค้นหาผิดพลาด: {e}"))
                return
            self.root.after(0, lambda: self.show_results(results))

        threading.Thread(target=worker, daemon=True).start()

    def show_results(self, results):
        self.search_results = results
        self._render_search_results()

    def _render_search_results(self):
        for widget in self.results_frame.winfo_children():
            widget.destroy()
        self._results_row_labels = {}
        self._results_selected_index = None
        for i, r in enumerate(self.search_results):
            row = tk.Frame(self.results_frame, bg="white")
            row.pack(fill="x")
            name_label = tk.Label(
                row, text=r["name"], font=("Tahoma", fs(10)), anchor="w", bg="white",
                padx=fs(4), pady=fs(3),
            )
            name_label.pack(side="left", fill="x", expand=True)
            name_label.bind("<Button-1>", lambda e, idx=i: self._select_search_result(idx))
            name_label.bind("<Double-Button-1>", lambda e, idx=i: self.on_add_from_search(idx))
            self._results_row_labels[i] = name_label
            tk.Button(
                row, text="🗑", fg="white", bg="#c0392b", font=("Tahoma", fs(8), "bold"), width=2,
                command=lambda idx=i: self.on_delete_from_db(idx),
            ).pack(side="left", padx=(fs(4), fs(2)))
        self._bind_mousewheel(self.results_frame, self._results_canvas)
        self._bind_mousewheel(self._results_canvas, self._results_canvas)

    def _select_search_result(self, index):
        """Highlight the clicked row blue (like a native Listbox selection)
        so there's visual feedback before the double-click that actually adds
        it - the highlight only covers the name label, not the delete button,
        so it never overlaps/tints the (red) trash icon."""
        prev = self._results_selected_index
        if prev is not None and prev in self._results_row_labels:
            self._results_row_labels[prev].config(bg="white", fg="black")
        label = self._results_row_labels.get(index)
        if label:
            label.config(bg="#0078d7", fg="white")
        self._results_selected_index = index

    # ---------------------------------------------------------------- selected-drug list

    def on_add_from_search(self, index):
        product = self.search_results[index]
        self.status_var.set(f"กำลังโหลดข้อมูล {product['name']}...")

        def worker():
            try:
                info = get_product_med_info(product["idproduct"])
            except Exception as e:
                self.root.after(0, lambda: self.status_var.set(f"เกิดข้อผิดพลาด: {e}"))
                return
            self.root.after(0, lambda: self.add_drug(product, info))

        threading.Thread(target=worker, daemon=True).start()

    def on_delete_from_db(self, index):
        product = self.search_results[index]
        if not messagebox.askyesno(
            "ยืนยัน", f"ลบ '{product['name']}' ออกจากฐานข้อมูลถาวรไหม?\n(ข้อมูลวิธีใช้ที่เคยกรอกไว้จะหายไปด้วย)",
        ):
            return

        def worker():
            try:
                storage.delete_template(product["idproduct"])
            except Exception as e:
                self.root.after(0, lambda: self.status_var.set(f"ลบไม่สำเร็จ: {e}"))
                return
            self.root.after(0, lambda: self._after_delete_from_db(product))

        threading.Thread(target=worker, daemon=True).start()

    def _after_delete_from_db(self, product):
        self.search_results = [r for r in self.search_results if r["idproduct"] != product["idproduct"]]
        self._render_search_results()
        self.status_var.set(f"ลบ '{product['name']}' ออกจากฐานข้อมูลแล้ว")

    def add_drug(self, product, info):
        # status: "db" = saved locally right now (green)
        #         "edited" = changed in this session but NOT saved yet (orange)
        #         "missing" = no dosing data yet (red) - this also covers a row
        #         that exists (e.g. from a bulk Excel import) but only has
        #         drug1 filled in, so it doesn't get misleadingly marked green
        has_data = has_dosing_data(info)
        if has_data:
            entry = {
                "idproduct": product["idproduct"],
                "drug1": info["drug1"] or product["name"],
                "drug2": info["drug2"],
                "note": info["note"],
                "qty": info["qty"],
                "unit": info["unit"] or "เม็ด",
                "per_day": info["per_day"],
                "every_hr": info["every_hr"],
                "meal": info["meal"] or "หลังอาหาร",
                "times": list(info["times"]),
                "extra_labels": list(info.get("extra_labels") or []),
                "usage_mode": info.get("usage_mode", "oral"),
                "barcode": info.get("barcode", ""),
                "status": "db",
                "print_qty": 1,
                "exp_date": info.get("exp_date", ""),
                "label_qty": info.get("label_qty", ""),
                "lang": "th",
                # Cached Grok translations of "note" - only valid as long as
                # "note" itself doesn't change (see collect_into_d, which
                # clears these if the pharmacist edits the Indication text).
                "note_en": info.get("note_en", ""),
                "note_mm": info.get("note_mm", ""),
            }
        else:
            entry = {
                "idproduct": product["idproduct"],
                "drug1": (info["drug1"] if info else None) or product["name"],
                "drug2": "", "note": "", "qty": "", "unit": "เม็ด",
                "per_day": "", "every_hr": "", "meal": "หลังอาหาร", "times": [],
                "extra_labels": [],
                "usage_mode": (info.get("usage_mode", "oral") if info else "oral"),
                "barcode": (info.get("barcode", "") if info else ""),
                "status": "missing",
                "print_qty": 1,
                "exp_date": (info.get("exp_date", "") if info else ""),
                "label_qty": (info.get("label_qty", "") if info else ""),
                "lang": "th",
                "note_en": (info.get("note_en", "") if info else ""),
                "note_mm": (info.get("note_mm", "") if info else ""),
            }
        self.selected_drugs.append(entry)
        self.refresh_selected_list()
        if has_data:
            self.status_var.set(f"เพิ่ม {entry['drug1']} แล้ว")
        else:
            self.status_var.set(f"เพิ่ม {entry['drug1']} แล้ว (ยังไม่มีข้อมูลวิธีกิน - ดับเบิลคลิกที่รายการเพื่อกรอก)")

    def add_new_drug(self):
        """+ เพิ่มยาใหม่ - there's no external product catalog here (unlike the
        POS-connected version), so a brand-new drug starts completely blank."""
        entry = {
            "idproduct": None, "drug1": "", "drug2": "", "note": "", "qty": "",
            "unit": "เม็ด", "per_day": "", "every_hr": "", "meal": "หลังอาหาร",
            "times": [], "extra_labels": [], "usage_mode": "oral", "barcode": "",
            "status": "missing", "print_qty": 1, "exp_date": "", "label_qty": "", "lang": "th",
            "note_en": "", "note_mm": "",
        }
        self.selected_drugs.append(entry)
        index = len(self.selected_drugs) - 1
        self.refresh_selected_list()
        self.open_edit_dialog(index, is_new=True)

    def open_import_excel_dialog(self):
        """Bulk-create blank drug templates (ชื่อการค้า + optional
        บาร์โค้ด/ชื่อสามัญ) from an Excel file - for a pharmacy migrating a
        list of drug names in from elsewhere. Never overwrites dosing info
        on a drug that's already saved.

        No column-name entry fields here on purpose - real-world exported
        files have wildly inconsistent/multi-row headers (see
        read_excel_drug_names_and_barcodes()'s docstring), so instead of
        asking the pharmacist to type an exact header string that's easy to
        typo, the file itself gets pre-edited (by whoever prepares it) to
        use the fixed col1/col2/col3 header names - always found instantly,
        no typing needed here at all. The auto-detect list still also tries
        several common Thai/English header names as a bonus, in case
        someone imports a file without col1/col2/col3."""
        win = tk.Toplevel(self.root)
        win.title("Import รายชื่อยาจาก Excel")
        win.geometry(f"{fs(440)}x{fs(320)}")
        win.transient(self.root)
        win.grab_set()

        pad = {"padx": fs(10), "pady": fs(4)}

        tk.Label(
            win, text="ก่อน import ให้เปิดไฟล์ Excel แล้วเปลี่ยนชื่อหัวตาราง (แถวแรก) ของ 3 คอลัมน์นี้ก่อน แล้ว Save ไฟล์:\n"
                       "• คอลัมน์ชื่อยา → เปลี่ยนเป็น col1\n"
                       "• คอลัมน์บาร์โค้ด (ไม่บังคับ) → เปลี่ยนเป็น col2\n"
                       "• คอลัมน์ชื่อสามัญทางยา (ไม่บังคับ) → เปลี่ยนเป็น col3\n"
                       "(พิมพ์ตัวเล็กตัวใหญ่ไม่มีผล) เสร็จแล้วค่อยเลือกไฟล์มา import ด้านล่าง ระบบจะหาคอลัมน์ให้เองอัตโนมัติ",
            font=("Tahoma", fs(9), "bold"), wraplength=fs(400), justify="left",
        ).pack(anchor="w", **pad)

        tk.Label(
            win, text="ยาที่ยังไม่มีในเครื่องจะถูกสร้างรายการเปล่าให้ (ยังไม่มีวิธีใช้ กรอกเพิ่มทีหลังได้) "
                       "ส่วนยาที่มีอยู่แล้วจะ**ไม่ถูกเขียนทับ** - ยกเว้นบาร์โค้ด/ชื่อสามัญที่จะอัปเดตให้ถ้าไฟล์มีค่าใหม่มา",
            font=("Tahoma", fs(8)), fg="#555", wraplength=fs(400), justify="left",
        ).pack(anchor="w", **pad)

        file_var = tk.StringVar(value="")
        file_row = tk.Frame(win)
        file_row.pack(fill="x", **pad)
        tk.Entry(file_row, textvariable=file_var, font=("Tahoma", fs(10)), state="readonly").pack(
            side="left", fill="x", expand=True)

        def choose_file():
            path = filedialog.askopenfilename(
                title="เลือกไฟล์ Excel", filetypes=[("Excel files", "*.xlsx")], parent=win,
            )
            if path:
                file_var.set(path)

        tk.Button(file_row, text="เลือกไฟล์...", font=("Tahoma", fs(9)), command=choose_file).pack(
            side="left", padx=(fs(4), 0))

        status_var = tk.StringVar(value="")
        tk.Label(win, textvariable=status_var, font=("Tahoma", fs(14), "bold"), fg="#a00",
                 wraplength=fs(400), justify="left").pack(**pad)

        def do_import():
            path = file_var.get().strip()
            if not path:
                status_var.set("กรุณาเลือกไฟล์ Excel ก่อน")
                return
            import_btn.config(state="disabled")
            status_var.set("กำลัง import...")

            def worker():
                try:
                    rows = read_excel_drug_names_and_barcodes(path)
                except Exception as e:
                    self.root.after(0, lambda: status_var.set(f"อ่านไฟล์ไม่สำเร็จ: {e}"))
                    self.root.after(0, lambda: import_btn.config(state="normal"))
                    return
                try:
                    created, updated_barcode, skipped_blank = storage.bulk_import_names_and_barcodes(rows)
                except Exception as e:
                    self.root.after(0, lambda: status_var.set(f"บันทึกไม่สำเร็จ: {e}"))
                    self.root.after(0, lambda: import_btn.config(state="normal"))
                    return

                def done():
                    messagebox.showinfo(
                        "สำเร็จ",
                        f"Import เสร็จแล้ว\nเพิ่มใหม่: {created} รายการ\n"
                        f"อัปเดตบาร์โค้ด (ยาที่มีอยู่แล้ว): {updated_barcode} รายการ\n"
                        f"ข้าม (ชื่อซ้ำ/ว่างในไฟล์): {skipped_blank} รายการ",
                        parent=win,
                    )
                    self.status_var.set(f"Import จาก Excel สำเร็จ: เพิ่มยาใหม่ {created} รายการ")
                    win.destroy()
                self.root.after(0, done)

            threading.Thread(target=worker, daemon=True).start()

        import_btn = tk.Button(
            win, text="📥 เริ่ม Import", font=("Tahoma", fs(10), "bold"),
            bg="#1a7a4a", fg="white", command=do_import,
        )
        import_btn.pack(pady=fs(10))

        win.lift()
        win.focus_force()

    def refresh_selected_list(self):
        for widget in self.list_frame.winfo_children():
            widget.destroy()

        marks = {"db": "✓", "edited": "✎", "missing": "✗", "override": "★"}
        colors = {"db": "#1a7a4a", "edited": "#c07a17", "missing": "#b03a2e", "override": "#b8960c"}

        for i, d in enumerate(self.selected_drugs):
            row = tk.Frame(self.list_frame)
            row.pack(fill="x", pady=fs(1))

            status = d.get("status", "missing")
            mark = marks[status]
            color = colors[status]
            name_label = tk.Label(
                row, text=f"{mark} {d['drug1'] or '(ยังไม่ตั้งชื่อ)'}", font=("Tahoma", fs(10)),
                bg=color, fg="white", anchor="w", padx=fs(6), pady=fs(5),
            )
            name_label.pack(side="left", fill="x", expand=True)
            name_label.bind("<Double-Button-1>", lambda e, idx=i: self.open_edit_dialog(idx))

            # exp_date/label_qty are remembered per-drug (drug_templates
            # columns) once the drug already has a saved template - a shop
            # dispenses from the same lot for a while, so auto-persisting on
            # every edit means the next time this drug is added, both fields
            # come back prefilled instead of typed fresh each time. A
            # brand-new/unsaved drug (idproduct None) just keeps the value
            # in-memory for this session, same as before - don't silently
            # create a half-filled DB row from typing only these 2 fields.
            def persist_exp_and_qty(idx):
                entry = self.selected_drugs[idx]
                idproduct = entry.get("idproduct")
                if not idproduct:
                    return

                def worker():
                    try:
                        save_product_med_info(idproduct, entry)
                    except Exception:
                        pass  # best-effort remember - never block typing on a DB hiccup

                threading.Thread(target=worker, daemon=True).start()

            tk.Label(row, text="EXP", font=("Tahoma", fs(10))).pack(side="left", padx=(fs(6), fs(2)))
            exp_var = tk.StringVar(value=str(d.get("exp_date", "")))

            def on_exp_change(*args, idx=i, var=exp_var):
                self.selected_drugs[idx]["exp_date"] = var.get().strip()
                persist_exp_and_qty(idx)

            exp_var.trace_add("write", on_exp_change)
            tk.Entry(row, textvariable=exp_var, width=6, font=("Tahoma", fs(10)), justify="center").pack(side="left", padx=fs(2))

            # #บนฉลาก - the quantity actually printed ON the label itself
            # (e.g. "#30" for 30 tablets dispensed). Left blank, the label
            # keeps the handwritten blank line as before (see
            # build_label_image). Distinct from #ฉลาก below, which is how
            # many physical label copies to print.
            tk.Label(row, text="#บนฉลาก", font=("Tahoma", fs(10))).pack(side="left", padx=(fs(6), fs(2)))
            label_qty_var = tk.StringVar(value=str(d.get("label_qty", "")))

            def on_label_qty_change(*args, idx=i, var=label_qty_var):
                self.selected_drugs[idx]["label_qty"] = var.get().strip()
                persist_exp_and_qty(idx)

            label_qty_var.trace_add("write", on_label_qty_change)
            tk.Entry(row, textvariable=label_qty_var, width=5, font=("Tahoma", fs(10)), justify="center").pack(side="left", padx=fs(2))

            # Per-label language, chosen at print time (not remembered per
            # drug) - a shop's walk-in customers are often mixed
            # nationality, so the right language varies visit to visit, not
            # by drug. Default "ไทย" every time. See LABEL_LANGS.
            lang_var = tk.StringVar(value=LABEL_LANG_NAMES.get(d.get("lang", "th"), "ไทย"))
            lang_names_to_code = {v: k for k, v in LABEL_LANG_NAMES.items()}

            def on_lang_change(event=None, idx=i, var=lang_var):
                self.selected_drugs[idx]["lang"] = lang_names_to_code.get(var.get(), "th")

            lang_combo = ttk.Combobox(
                row, textvariable=lang_var, values=list(LABEL_LANG_NAMES.values()),
                state="readonly", font=("Tahoma", fs(9)), width=7,
            )
            lang_combo.pack(side="left", padx=(fs(6), fs(2)))
            lang_combo.bind("<<ComboboxSelected>>", on_lang_change)

            tk.Label(row, text="#ฉลาก", font=("Tahoma", fs(10))).pack(side="left", padx=(fs(6), fs(2)))
            qty_var = tk.StringVar(value=str(d.get("print_qty", 1)))

            def on_qty_change(*args, idx=i, var=qty_var):
                val = var.get().strip()
                if val.isdigit() and int(val) > 0:
                    self.selected_drugs[idx]["print_qty"] = int(val)

            qty_var.trace_add("write", on_qty_change)
            tk.Entry(row, textvariable=qty_var, width=3, font=("Tahoma", fs(10)), justify="center").pack(side="left", padx=fs(2))
            tk.Label(row, text="แผ่น", font=("Tahoma", fs(10))).pack(side="left", padx=(fs(2), fs(4)))

            tk.Button(
                row, text="👁 Preview", font=("Tahoma", fs(9)),
                command=lambda idx=i: self.preview_label(idx),
            ).pack(side="left", padx=(fs(4), 0))

            tk.Button(
                row, text="✕", fg="white", bg="#555555", font=("Tahoma", fs(9), "bold"),
                width=2, command=lambda idx=i: self.remove_drug(idx),
            ).pack(side="left", padx=(fs(4), 0))

        self._bind_mousewheel(self.list_frame, self._list_canvas)
        self._bind_mousewheel(self._list_canvas, self._list_canvas)

    def remove_drug(self, index):
        del self.selected_drugs[index]
        self.refresh_selected_list()

    def clear_all_drugs(self):
        if not self.selected_drugs:
            return
        if not messagebox.askyesno("ยืนยัน", f"ล้างรายการยาทั้งหมด ({len(self.selected_drugs)} รายการ) ใช่ไหม?"):
            return
        self.selected_drugs = []
        self.refresh_selected_list()
        self.status_var.set("ล้างรายการยาทั้งหมดแล้ว")

    def clear_selected_customer(self):
        self.selected_customer_var.set(self.NO_CUSTOMER_TEXT)
        self._queue_patient_name = None
        self._queue_patient_phone = None
        self._queue_patient_id = None

    def _get_allergy_drug_name(self, patient_id):
        """If patient_id has exactly one recorded drug allergy, return its
        name so it can be printed directly on the label instead of the
        generic "ดูแฟ้ม" notice. Multiple (or zero/unknown) allergies return
        "" and the label falls back to "ดูแฟ้ม" as before."""
        if not patient_id:
            return ""
        try:
            patient = storage.get_patient(patient_id)
        except Exception:
            patient = None
        if not patient:
            return ""
        items = split_allergy_items(patient.get("allergy_note") or "")
        return items[0] if len(items) == 1 else ""

    def preview_label(self, index):
        d = self.selected_drugs[index]
        lang = d.get("lang", "th")
        note = (d.get("note") or "").strip()
        cache_key = f"note_{lang}"
        # Indication (note) is free-typed and never auto-translated except
        # here, on demand: if this drug has Thai note text and the label
        # language isn't Thai, and we don't already have a cached Grok
        # translation for it, fetch one now (threaded - Grok can take a
        # few seconds and must never freeze the window) before opening the
        # preview, so the preview always shows what will actually print.
        if lang in ("en", "mm") and note and not d.get(cache_key):
            self.status_var.set("กำลังแปล Indication ด้วย Grok...")

            def worker():
                translated, _err = translate_note_via_grok(note, lang)
                def done():
                    self.status_var.set("")
                    if translated:
                        d[cache_key] = translated
                        idproduct = d.get("idproduct")
                        if idproduct:
                            storage.save_note_translation(idproduct, lang, translated)
                    # On failure (no key configured, network error, etc.) -
                    # fall back silently to the original Thai text. This is
                    # a rarely-used convenience feature, not worth a popup
                    # every time; the preview itself already makes it
                    # obvious the note stayed Thai.
                    self._show_label_preview(index)
                self.root.after(0, done)

            threading.Thread(target=worker, daemon=True).start()
            return
        self._show_label_preview(index)

    def _show_label_preview(self, index):
        d = self.selected_drugs[index]
        lang = d.get("lang", "th")
        data = dict(d)
        if lang in ("en", "mm"):
            translated = d.get(f"note_{lang}")
            if translated:
                data["note"] = translated
        data["patient_name"] = "(ชื่อผู้ป่วย)"
        data["has_allergy"] = self.allergy_var.get()
        if data["has_allergy"]:
            data["allergy_drug_name"] = self._get_allergy_drug_name(self._queue_patient_id)
        settings = app_settings.load_settings()
        img = build_label_image(data, settings)

        win = tk.Toplevel(self.root)
        win.title(f"ตัวอย่างฉลาก - {d['drug1']}")
        win.transient(self.root)

        scale = 1.6
        preview_img = img.resize((int(img.width * scale), int(img.height * scale)), Image.LANCZOS)
        photo = ImageTk.PhotoImage(preview_img)

        label = tk.Label(win, image=photo, bd=2, relief="solid")
        label.image = photo  # keep a reference so it isn't garbage-collected
        label.pack(padx=fs(10), pady=fs(10))

        tk.Label(win, text="ชื่อผู้ป่วยจริงจะถูกใส่ตอนกดยืนยันพิมพ์ - รูปนี้เป็นตัวอย่างเท่านั้น",
                 font=("Tahoma", fs(9)), fg="#777").pack(pady=(0, fs(8)))
        tk.Button(win, text="ปิด", font=("Tahoma", fs(10)), command=win.destroy).pack(pady=(0, fs(10)))

        win.lift()
        win.focus_force()

    # ---------------------------------------------------------------- favorites

    def refresh_fav_buttons(self):
        for widget in self.fav_button_frame.winfo_children():
            widget.destroy()

        term = self.fav_search_var.get().strip().lower()
        if term:
            names = [n for n in self.favorites if term in n.lower()]
            names.sort(key=lambda n: (0 if n.lower().startswith(term) else 1, n.lower()))
        else:
            names = sorted(self.favorites.keys())

        if not names:
            tk.Label(self.fav_button_frame, text="(ไม่มี Favorite)", font=("Tahoma", fs(9)),
                     bg="#f0f0f0", fg="#888").pack(pady=fs(8))
        else:
            for name in names:
                row = tk.Frame(self.fav_button_frame, bg="#f0f0f0")
                row.pack(fill="x", pady=fs(2))
                tk.Button(
                    row, text=name, font=("Tahoma", fs(9)), wraplength=fs(140), justify="left", anchor="w",
                    bg="#ffffff", relief="raised", command=lambda n=name: self.on_load_favorite(n),
                ).pack(side="left", fill="x", expand=True)
                tk.Button(
                    row, text="💾", font=("Tahoma", fs(8)), width=2,
                    command=lambda n=name: self.on_overwrite_favorite(n),
                ).pack(side="left", padx=(fs(2), 0))
                tk.Button(
                    row, text="✕", font=("Tahoma", fs(8), "bold"), fg="white", bg="#888888", width=2,
                    command=lambda n=name: self.on_delete_favorite(n),
                ).pack(side="left", padx=(fs(2), 0))

        self._bind_mousewheel(self.fav_button_frame, self._fav_canvas)
        self._bind_mousewheel(self._fav_canvas, self._fav_canvas)

    def on_load_favorite(self, name):
        if name not in self.favorites:
            return
        saved_entries = [dict(e) for e in self.favorites[name]]
        self.status_var.set(f"กำลังโหลด Favorite '{name}'...")

        def worker():
            loaded = []
            for entry in saved_entries:
                # An override entry was deliberately frozen at save-favorite
                # time (see on_save_favorite) specifically so it does NOT get
                # clobbered by whatever the DB template currently says - skip
                # the refresh entirely and trust the favorite's own copy.
                if entry.get("override"):
                    entry["status"] = "override"
                    entry.setdefault("usage_mode", "oral")
                    entry.setdefault("print_qty", 1)
                    loaded.append(entry)
                    continue
                info = None
                if entry.get("idproduct") is not None:
                    try:
                        info = get_product_med_info(entry["idproduct"])
                    except Exception:
                        info = None
                # Always re-check storage instead of trusting the status saved
                # into the favorite at save-time - a drug that was "missing"
                # back then may have gotten its info filled in since, and
                # should show green immediately, not stay stale red.
                if has_dosing_data(info):
                    entry["drug1"] = info["drug1"] or entry.get("drug1", "")
                    entry["drug2"] = info["drug2"]
                    entry["note"] = info["note"]
                    entry["qty"] = info["qty"]
                    entry["unit"] = info["unit"] or "เม็ด"
                    entry["per_day"] = info["per_day"]
                    entry["every_hr"] = info["every_hr"]
                    entry["meal"] = info["meal"] or "หลังอาหาร"
                    entry["times"] = list(info["times"])
                    entry["extra_labels"] = list(info.get("extra_labels") or [])
                    entry["usage_mode"] = info.get("usage_mode", "oral")
                    entry["barcode"] = info.get("barcode", "")
                    entry["status"] = "db"
                else:
                    entry["status"] = "missing"
                entry.setdefault("usage_mode", "oral")
                entry.setdefault("print_qty", 1)
                loaded.append(entry)

            def apply():
                self.selected_drugs.extend(loaded)
                self.refresh_selected_list()
                self.status_var.set(f"โหลด Favorite '{name}' แล้ว ({len(loaded)} รายการยา)")

            self.root.after(0, apply)

        threading.Thread(target=worker, daemon=True).start()

    def on_delete_favorite(self, name):
        if name not in self.favorites:
            return
        if not messagebox.askyesno("ยืนยัน", f"ลบ Favorite '{name}' ใช่ไหม?"):
            return
        del self.favorites[name]
        save_favorites(self.favorites)
        self.refresh_fav_buttons()
        self.status_var.set(f"ลบ Favorite '{name}' แล้ว")

    def on_save_favorite(self):
        if not self.selected_drugs:
            messagebox.showwarning("แจ้งเตือน", "กรุณาเลือกยาในรายการก่อนบันทึกเป็น Favorite")
            return
        name = simpledialog.askstring("บันทึก Favorite", "ตั้งชื่อชุดยานี้:", parent=self.root)
        if not name:
            return
        name = name.strip()
        if not name:
            return
        if name in self.favorites:
            if not messagebox.askyesno("แจ้งเตือน", f"มี Favorite ชื่อ '{name}' อยู่แล้ว ต้องการเขียนทับไหม?"):
                return
        # A drug edited away from its DB default (orange "edited") gets
        # frozen into this favorite as an "override" - on_load_favorite()
        # will skip re-fetching it from the DB template next time, so a
        # dose customized specifically for this combo doesn't get silently
        # thrown away the way it used to (every load unconditionally
        # overwrote with the current global template). A drug that matches
        # the DB (or has no data at all) has nothing to protect, so it stays
        # non-override and keeps auto-refreshing from the DB as before.
        snapshot = []
        for d in self.selected_drugs:
            entry = dict(d)
            entry["override"] = d.get("status") == "edited"
            snapshot.append(entry)
        self.favorites[name] = snapshot
        save_favorites(self.favorites)
        self.refresh_fav_buttons()
        self.status_var.set(f"บันทึก Favorite '{name}' แล้ว ({len(self.selected_drugs)} รายการยา)")

    def on_overwrite_favorite(self, name):
        if name not in self.favorites:
            return
        if not self.selected_drugs:
            messagebox.showwarning("แจ้งเตือน", "กรุณาเลือกยาในรายการก่อนบันทึกทับ Favorite")
            return
        if not messagebox.askyesno("ยืนยัน", f"บันทึกทับ Favorite '{name}' ด้วยรายการปัจจุบันใช่ไหม?"):
            return
        snapshot = []
        for d in self.selected_drugs:
            entry = dict(d)
            entry["override"] = d.get("status") == "edited"
            snapshot.append(entry)
        self.favorites[name] = snapshot
        save_favorites(self.favorites)
        self.refresh_fav_buttons()
        self.status_var.set(f"บันทึกทับ Favorite '{name}' แล้ว ({len(self.selected_drugs)} รายการยา)")

    def _sync_copied_target(self, idproduct, drug):
        """After a copy-to save succeeds, refresh any matching row already in
        the print queue so its color/data reflect the new data without
        needing to remove and re-add it."""
        changed = False
        for entry in self.selected_drugs:
            if entry["idproduct"] == idproduct:
                entry.update({
                    "drug2": drug["drug2"], "note": drug["note"], "qty": drug["qty"],
                    "unit": drug["unit"], "per_day": drug["per_day"], "every_hr": drug["every_hr"],
                    "meal": drug["meal"], "times": list(drug["times"]),
                    "extra_labels": list(drug.get("extra_labels") or []),
                    "status": "db",
                })
                changed = True
        if changed:
            self.refresh_selected_list()

    def open_copy_to_dialog(self, parent_win, current_idproduct, get_current_values):
        """Popup to search for one or more other drugs and push the dosing
        info currently in the edit dialog out to them (everything except
        ชื่อการค้า), saving straight to local storage for each."""
        win = tk.Toplevel(parent_win)
        win.title("Copy ข้อมูลนี้ไปยาอื่น")
        win.geometry(f"{fs(420)}x{fs(460)}")
        win.transient(parent_win)
        win.grab_set()

        tk.Label(
            win, text="ค้นหายาแล้วดับเบิลคลิกเพื่อเพิ่มเข้ารายการที่จะ copy ไปให้ (เลือกได้หลายรายการ)",
            font=("Tahoma", fs(9), "bold"), wraplength=fs(380), justify="left",
        ).pack(anchor="w", padx=fs(10), pady=(fs(10), fs(4)))

        search_var = tk.StringVar()
        tk.Entry(win, textvariable=search_var, font=("Tahoma", fs(11))).pack(fill="x", padx=fs(10), pady=(0, fs(4)))
        results_listbox = tk.Listbox(win, height=6, font=("Tahoma", fs(10)))
        results_listbox.pack(fill="x", padx=fs(10), pady=(0, fs(6)))

        tk.Label(win, text="รายการที่จะ copy ไปให้ (ดับเบิลคลิกเพื่อลบออก) - เขียว = มีข้อมูลอยู่แล้ว (จะถูกเขียนทับ), แดง = ยังไม่มีข้อมูล:",
                 font=("Tahoma", fs(9), "bold"), wraplength=fs(380), justify="left").pack(anchor="w", padx=fs(10))
        targets_listbox = tk.Listbox(win, height=6, font=("Tahoma", fs(10)))
        targets_listbox.pack(fill="both", expand=True, padx=fs(10), pady=(0, fs(6)))

        status_var = tk.StringVar(value="")
        tk.Label(win, textvariable=status_var, font=("Tahoma", fs(9)), fg="#a00",
                 wraplength=fs(380), justify="left").pack(padx=fs(10), pady=(0, fs(4)))

        results = []
        targets = []
        after_id = [None]

        def do_search():
            term = search_var.get().strip()
            results_listbox.delete(0, tk.END)
            results.clear()
            if not term:
                return

            def worker():
                try:
                    found = search_products(term)
                except Exception as e:
                    win.after(0, lambda: status_var.set(f"ค้นหาผิดพลาด: {e}"))
                    return

                def update():
                    results.extend(found)
                    for r in found:
                        results_listbox.insert(tk.END, r["name"])
                win.after(0, update)

            threading.Thread(target=worker, daemon=True).start()

        def on_search_change(*a):
            if after_id[0]:
                win.after_cancel(after_id[0])
            after_id[0] = win.after(300, do_search)

        search_var.trace_add("write", on_search_change)

        def on_add_target(event=None):
            sel = results_listbox.curselection()
            if not sel:
                return
            product = results[sel[0]]
            if product["idproduct"] == current_idproduct:
                status_var.set("ไม่สามารถ copy ไปยาตัวเองได้")
                return
            if any(t["idproduct"] == product["idproduct"] for t in targets):
                status_var.set(f"'{product['name']}' อยู่ในรายการแล้ว")
                return
            targets.append(product)
            targets_listbox.insert(tk.END, f"...  {product['name']}")
            targets_listbox.itemconfig(tk.END, fg="#888")
            status_var.set("")

            def worker():
                try:
                    info = get_product_med_info(product["idproduct"])
                except Exception:
                    info = None

                def update():
                    if product not in targets:
                        return  # ลบออกไปก่อนที่จะเช็คสถานะเสร็จ
                    live_idx = targets.index(product)
                    mark, color = ("✓", "#1a7a4a") if has_dosing_data(info) else ("✗", "#b03a2e")
                    targets_listbox.delete(live_idx)
                    targets_listbox.insert(live_idx, f"{mark}  {product['name']}")
                    targets_listbox.itemconfig(live_idx, fg=color)
                win.after(0, update)

            threading.Thread(target=worker, daemon=True).start()

        def on_remove_target(event=None):
            sel = targets_listbox.curselection()
            if not sel:
                return
            idx = sel[0]
            targets_listbox.delete(idx)
            del targets[idx]

        results_listbox.bind("<Double-Button-1>", on_add_target)
        targets_listbox.bind("<Double-Button-1>", on_remove_target)

        def on_confirm():
            if not targets:
                messagebox.showwarning("แจ้งเตือน", "กรุณาเพิ่มยาที่จะ copy ไปให้อย่างน้อย 1 รายการ", parent=win)
                return
            if not messagebox.askyesno(
                "ยืนยัน",
                f"Copy ข้อมูลนี้ (ยกเว้นชื่อการค้า) ไปยัง {len(targets)} รายการ และบันทึกเลยหรือไม่?",
                parent=win,
            ):
                return
            confirm_btn.config(state="disabled")
            values = get_current_values()
            target_list = list(targets)

            def worker():
                errors = []
                for t in target_list:
                    try:
                        drug = dict(values)
                        drug["drug1"] = t["name"]
                        save_product_med_info(t["idproduct"], drug)
                        self.root.after(0, lambda t=t, drug=drug: self._sync_copied_target(t["idproduct"], drug))
                    except Exception as e:
                        errors.append(f"{t['name']}: {e}")

                def done():
                    confirm_btn.config(state="normal")
                    if errors:
                        messagebox.showerror("บันทึกไม่สำเร็จบางรายการ", "\n".join(errors), parent=win)
                    else:
                        messagebox.showinfo(
                            "สำเร็จ", f"Copy ไปยัง {len(target_list)} รายการ และบันทึกแล้ว", parent=win)
                        win.destroy()
                self.root.after(0, done)

            threading.Thread(target=worker, daemon=True).start()

        confirm_btn = tk.Button(
            win, text="✅ Confirm copy", font=("Tahoma", fs(10), "bold"),
            bg="#1a7a4a", fg="white", command=on_confirm,
        )
        confirm_btn.pack(pady=(0, fs(10)))

        win.lift()
        win.focus_force()

    def open_copy_from_dialog(self, parent_win, target_vars):
        """Popup to search for another drug and copy its dosing info (everything
        except ชื่อการค้า) into the edit dialog currently open - for drugs that
        are the same medicine under a different brand name."""
        win = tk.Toplevel(parent_win)
        win.title("Copy ข้อมูลจากยาอื่น")
        win.geometry(f"{fs(420)}x{fs(340)}")
        win.transient(parent_win)
        win.grab_set()

        tk.Label(
            win, text="ค้นหายาที่จะ copy ข้อมูลมา (ต้องเป็นยาที่มีข้อมูลครบแล้ว - ขึ้นเขียวในรายการ)",
            font=("Tahoma", fs(9), "bold"), wraplength=fs(380), justify="left",
        ).pack(anchor="w", padx=fs(10), pady=(fs(10), fs(4)))

        search_var = tk.StringVar()
        tk.Entry(win, textvariable=search_var, font=("Tahoma", fs(11))).pack(fill="x", padx=fs(10), pady=(0, fs(4)))
        listbox = tk.Listbox(win, height=8, font=("Tahoma", fs(10)))
        listbox.pack(fill="both", expand=True, padx=fs(10), pady=(0, fs(6)))

        status_var = tk.StringVar(value="")
        tk.Label(win, textvariable=status_var, font=("Tahoma", fs(9)), fg="#a00",
                 wraplength=fs(380), justify="left").pack(padx=fs(10), pady=(0, fs(8)))

        results = []
        after_id = [None]

        def do_search():
            term = search_var.get().strip()
            listbox.delete(0, tk.END)
            results.clear()
            if not term:
                return

            def worker():
                try:
                    found = search_products(term)
                except Exception as e:
                    win.after(0, lambda: status_var.set(f"ค้นหาผิดพลาด: {e}"))
                    return

                def update():
                    results.extend(found)
                    for r in found:
                        listbox.insert(tk.END, r["name"])
                win.after(0, update)

            threading.Thread(target=worker, daemon=True).start()

        def on_search_change(*a):
            if after_id[0]:
                win.after_cancel(after_id[0])
            after_id[0] = win.after(300, do_search)

        search_var.trace_add("write", on_search_change)

        def apply_copy(product, info):
            if not has_dosing_data(info):
                status_var.set(f"'{product['name']}' ยังไม่มีข้อมูลในระบบ (ไม่ใช่สีเขียว) - copy ไม่ได้")
                return
            target_vars["drug2_var"].set(info["drug2"])
            target_vars["note_var"].set(info["note"])
            target_vars["qty_var"].set(info["qty"])
            target_vars["unit_var"].set(info["unit"] or "เม็ด")
            target_vars["per_day_var"].set(info["per_day"])
            target_vars["every_hr_var"].set(info["every_hr"])
            target_vars["every_hr_enabled_var"].set(bool(info["every_hr"]))
            meal_value = info["meal"] or "หลังอาหาร"
            target_vars["meal_var"].set(meal_value)
            target_vars["meal_display_var"].set(MEAL_VALUE_TO_DISPLAY.get(meal_value, MEAL_OPTIONS_DISPLAY[-1]))
            for t in TIME_OPTIONS:
                target_vars["time_vars"][t].set(t in info["times"])
            usage_mode = info.get("usage_mode", "oral")
            target_vars["mode_var"].set(usage_mode)
            target_vars["mode_display_var"].set(USAGE_MODE_LABELS[usage_mode])
            target_vars["render_dose_fields"]()
            target_vars["render_extra_fields"]()
            # render_extra_fields() just rebuilt extra_vars for the new mode's
            # option set - only now can checked state be set against it.
            extra_info = info.get("extra_labels") or []
            for opt, v in target_vars["extra_vars"].items():
                v.set(opt in extra_info)
            custom_candidate = next((e for e in extra_info if e not in target_vars["extra_vars"]), "")
            target_vars["custom_var"].set(bool(custom_candidate))
            target_vars["custom_text_var"].set(custom_candidate)
            win.destroy()
            messagebox.showinfo(
                "สำเร็จ", f"Copy ข้อมูลจาก '{product['name']}' มาแล้ว (ยกเว้นชื่อการค้า)", parent=parent_win,
            )

        def on_pick(event=None):
            sel = listbox.curselection()
            if not sel:
                return
            product = results[sel[0]]
            status_var.set("กำลังโหลดข้อมูล...")

            def worker():
                try:
                    info = get_product_med_info(product["idproduct"])
                except Exception as e:
                    win.after(0, lambda: status_var.set(f"เกิดข้อผิดพลาด: {e}"))
                    return
                win.after(0, lambda: apply_copy(product, info))

            threading.Thread(target=worker, daemon=True).start()

        listbox.bind("<Double-Button-1>", on_pick)
        win.lift()
        win.focus_force()

    def open_edit_dialog(self, index, is_new=False):
        d = self.selected_drugs[index]
        win = tk.Toplevel(self.root)
        win.title(f"แก้ไขข้อมูลยา - {d['drug1'] or '(ยาใหม่)'}")
        win.geometry(f"{fs(590)}x{fs(720)}")  # 440 * 4/3 - room for a 3-column ฉลากเสริม grid
        win.transient(self.root)
        win.grab_set()

        pad = {"padx": fs(10), "pady": fs(4)}

        tk.Label(win, text="ชื่อการค้า *", font=("Tahoma", fs(10), "bold")).pack(anchor="w", **pad)
        name_row = tk.Frame(win)
        name_row.pack(fill="x", **pad)
        drug1_var = tk.StringVar(value=d["drug1"])
        tk.Entry(name_row, textvariable=drug1_var, font=("Tahoma", fs(11))).pack(side="left", fill="x", expand=True)
        tk.Label(name_row, text="บาร์โค้ด:", font=("Tahoma", fs(9))).pack(side="left", padx=(fs(6), fs(2)))
        barcode_var = tk.StringVar(value=d.get("barcode", ""))
        tk.Entry(name_row, textvariable=barcode_var, font=("Tahoma", fs(11)), width=14).pack(side="left")

        tk.Label(win, text="ประเภทการใช้ยา", font=("Tahoma", fs(10), "bold")).pack(anchor="w", **pad)
        mode_var = tk.StringVar(value=d.get("usage_mode", "oral"))
        mode_label_to_key = {v: k for k, v in USAGE_MODE_LABELS.items()}
        mode_display_var = tk.StringVar(value=USAGE_MODE_LABELS[mode_var.get()])
        mode_combo = ttk.Combobox(
            win, textvariable=mode_display_var, values=list(USAGE_MODE_LABELS.values()),
            state="readonly", font=("Tahoma", fs(12)),
        )
        mode_combo.pack(fill="x", **pad)

        def get_current_values():
            return {
                "drug2": drug2_var.get().strip(),
                "note": note_var.get().strip(),
                "qty": qty_var.get().strip(),
                "unit": unit_var.get().strip(),
                "per_day": per_day_var.get().strip(),
                "every_hr": every_hr_var.get().strip() if every_hr_enabled_var.get() else "",
                "meal": meal_var.get().strip(),
                "times": [t for t, v in time_vars.items() if v.get()],
                "extra_labels": collect_extra_labels(),
                "usage_mode": mode_var.get(),
            }

        copy_btn_row = tk.Frame(win)
        copy_btn_row.pack(fill="x", **pad)
        tk.Button(
            copy_btn_row, text="📋 Copy จากยาอื่น", font=("Tahoma", fs(9), "bold"),
            bg="#5a5a9a", fg="white", command=lambda: self.open_copy_from_dialog(win, {
                "drug2_var": drug2_var, "note_var": note_var, "qty_var": qty_var,
                "unit_var": unit_var, "per_day_var": per_day_var, "every_hr_var": every_hr_var,
                "every_hr_enabled_var": every_hr_enabled_var,
                "meal_var": meal_var, "meal_display_var": meal_display_var,
                "time_vars": time_vars, "extra_vars": extra_vars,
                "custom_var": custom_var, "custom_text_var": custom_text_var,
                "mode_var": mode_var, "mode_display_var": mode_display_var,
                "render_dose_fields": lambda: render_dose_fields(),
                "render_extra_fields": lambda: render_extra_fields(),
            }),
        ).pack(side="left", fill="x", expand=True, padx=(0, fs(4)))
        tk.Button(
            copy_btn_row, text="📤 Copy ไปยาอื่น (หลายรายการ)", font=("Tahoma", fs(9), "bold"),
            bg="#5a5a9a", fg="white",
            command=lambda: self.open_copy_to_dialog(win, d["idproduct"], get_current_values),
        ).pack(side="left", fill="x", expand=True)

        tk.Label(win, text="ชื่อยาสามัญ", font=("Tahoma", fs(10), "bold")).pack(anchor="w", **pad)
        drug2_var = tk.StringVar(value=d["drug2"])
        tk.Entry(win, textvariable=drug2_var, font=("Tahoma", fs(11))).pack(fill="x", **pad)

        tk.Label(win, text="Indication (สรรพคุณ)", font=("Tahoma", fs(10), "bold")).pack(anchor="w", **pad)
        note_var = tk.StringVar(value=d["note"])
        tk.Entry(win, textvariable=note_var, font=("Tahoma", fs(11))).pack(fill="x", **pad)

        # These StringVars persist across mode switches (widgets get rebuilt,
        # values don't) - collect_into_d() below decides which ones actually
        # matter for the final saved mode.
        qty_var = tk.StringVar(value=d["qty"])
        unit_var = tk.StringVar(value=d["unit"] or "เม็ด")
        per_day_var = tk.StringVar(value=d["per_day"])
        every_hr_var = tk.StringVar(value=d["every_hr"] or "4")
        every_hr_enabled_var = tk.BooleanVar(value=bool(d["every_hr"]))
        meal_var = tk.StringVar(value=d["meal"] or "หลังอาหาร")
        meal_display_var = tk.StringVar(value=MEAL_VALUE_TO_DISPLAY.get(meal_var.get(), MEAL_OPTIONS_DISPLAY[-1]))

        dose_section = tk.Frame(win)
        dose_section.pack(fill="x")
        dose_font = ("Tahoma", fs(10))

        def render_dose_fields():
            for w in dose_section.winfo_children():
                w.destroy()
            mode = mode_var.get()
            dose_frame = tk.Frame(dose_section)
            dose_frame.pack(fill="x", **pad)
            if mode == "topical":
                tk.Label(dose_frame, text="ทาบางๆ วันละ", font=dose_font).grid(row=0, column=0)
                tk.Entry(dose_frame, textvariable=per_day_var, width=5, font=dose_font).grid(row=0, column=1, padx=fs(4))
                tk.Label(dose_frame, text="ครั้ง", font=dose_font).grid(row=0, column=2)
            elif mode == "drops":
                tk.Label(dose_frame, text="หยอดครั้งละ", font=dose_font).grid(row=0, column=0)
                tk.Entry(dose_frame, textvariable=qty_var, width=5, font=dose_font).grid(row=0, column=1, padx=fs(4))
                tk.Label(dose_frame, text="หยด   วันละ", font=dose_font).grid(row=0, column=2)
                tk.Entry(dose_frame, textvariable=per_day_var, width=5, font=dose_font).grid(row=0, column=3, padx=fs(4))
                tk.Label(dose_frame, text="ครั้ง", font=dose_font).grid(row=0, column=4)
            else:  # oral
                tk.Label(dose_frame, text="ครั้งละ", font=dose_font).grid(row=0, column=0)
                tk.Entry(dose_frame, textvariable=qty_var, width=5, font=dose_font).grid(row=0, column=1, padx=fs(4))
                unit_entry_frame = tk.Frame(dose_frame)
                unit_entry_frame.grid(row=0, column=2, padx=fs(4))
                tk.Entry(unit_entry_frame, textvariable=unit_var, width=6, font=dose_font).pack(side="left")
                unit_menu_btn = tk.Menubutton(unit_entry_frame, text="▾", font=dose_font, relief="raised")
                unit_menu = tk.Menu(unit_menu_btn, tearoff=0)
                for _opt in UNIT_OPTIONS:
                    unit_menu.add_command(label=_opt, command=lambda o=_opt: unit_var.set(o))
                unit_menu_btn.config(menu=unit_menu)
                unit_menu_btn.pack(side="left")
                tk.Label(dose_frame, text="วันละ", font=dose_font).grid(row=0, column=3)
                tk.Entry(dose_frame, textvariable=per_day_var, width=5, font=dose_font).grid(row=0, column=4, padx=fs(4))
                tk.Label(dose_frame, text="ครั้ง", font=dose_font).grid(row=0, column=5)

                tk.Label(dose_section, text="ก่อน/หลังอาหาร", font=("Tahoma", fs(10), "bold")).pack(anchor="w", **pad)
                meal_combo = ttk.Combobox(
                    dose_section, textvariable=meal_display_var, values=MEAL_OPTIONS_DISPLAY, state="readonly",
                    font=("Tahoma", fs(12)),
                )
                meal_combo.pack(fill="x", **pad)
                meal_combo.bind(
                    "<<ComboboxSelected>>",
                    lambda e: meal_var.set(MEAL_DISPLAY_TO_VALUE[meal_display_var.get()]),
                )

        tk.Label(win, text="เวลา (ติ๊กได้หลายอัน)", font=("Tahoma", fs(10), "bold")).pack(anchor="w", **pad)
        time_frame = tk.Frame(win)
        time_frame.pack(fill="x", **pad)
        time_vars = {}
        for t in TIME_OPTIONS:
            v = tk.BooleanVar(value=t in d["times"])
            time_vars[t] = v
            tk.Checkbutton(time_frame, text=t, variable=v, font=("Tahoma", fs(10))).pack(side="left", padx=fs(4))
        tk.Checkbutton(
            time_frame, text="หรือทุก", variable=every_hr_enabled_var, font=("Tahoma", fs(10)),
        ).pack(side="left", padx=(fs(8), 0))
        tk.Entry(time_frame, textvariable=every_hr_var, width=4, font=("Tahoma", fs(10))).pack(side="left", padx=fs(2))
        tk.Label(time_frame, text="ชม.", font=("Tahoma", fs(10))).pack(side="left")

        tk.Label(win, text=f"ฉลากเสริม (เลือกได้สูงสุด {MAX_EXTRA_LABELS} ข้อ)",
                 font=("Tahoma", fs(10), "bold")).pack(anchor="w", **pad)
        extra_section = tk.Frame(win)
        extra_section.pack(fill="x", **pad)
        extra_vars = {}
        # Free-text extra label ("อื่นๆ (พิมพ์เอง)") - a checkbox + entry pair
        # that isn't tied to any one usage mode's preset list, so these two
        # Variables are created once (not per render_extra_fields() call like
        # extra_vars is) and just get re-attached to fresh widgets each time,
        # so typed text survives a mode switch instead of getting wiped.
        custom_var = tk.BooleanVar(value=False)
        custom_text_var = tk.StringVar(value="")
        _custom_len_vcmd = (win.register(lambda P: len(P) <= MAX_CUSTOM_EXTRA_LABEL_CHARS), "%P")

        def _init_custom_from_d():
            all_presets = {opt for opts in EXTRA_LABEL_OPTIONS_BY_MODE.values() for opt in opts}
            existing_extra = d.get("extra_labels") or []
            custom_existing = next((e for e in existing_extra if e not in all_presets), "")
            custom_var.set(bool(custom_existing))
            custom_text_var.set(custom_existing)

        _init_custom_from_d()

        def checked_extra_count():
            return sum(1 for v in extra_vars.values() if v.get()) + (1 if custom_var.get() else 0)

        def collect_extra_labels():
            extra = [opt for opt, v in extra_vars.items() if v.get()]
            custom_text = custom_text_var.get().strip()
            if custom_var.get() and custom_text:
                extra.append(custom_text)
            return extra

        def on_extra_toggle(changed_key):
            if checked_extra_count() > MAX_EXTRA_LABELS:
                if changed_key == "__custom__":
                    custom_var.set(False)
                else:
                    extra_vars[changed_key].set(False)
                messagebox.showwarning("แจ้งเตือน", f"เลือกฉลากเสริมได้สูงสุด {MAX_EXTRA_LABELS} ข้อ", parent=win)

        EXTRA_GRID_COLUMNS = 3

        def render_extra_fields():
            for w in extra_section.winfo_children():
                w.destroy()
            extra_vars.clear()
            options = EXTRA_LABEL_OPTIONS_BY_MODE.get(mode_var.get(), EXTRA_LABEL_OPTIONS_BY_MODE["oral"])
            existing_extra = d.get("extra_labels") or []
            for col in range(EXTRA_GRID_COLUMNS):
                extra_section.grid_columnconfigure(col, weight=1)
            for i, opt in enumerate(options):
                v = tk.BooleanVar(value=opt in existing_extra)
                extra_vars[opt] = v
                row, col = divmod(i, EXTRA_GRID_COLUMNS)
                tk.Checkbutton(
                    extra_section, text=opt, variable=v, font=("Tahoma", fs(9)),
                    anchor="w", justify="left", wraplength=fs(150),
                    command=lambda o=opt: on_extra_toggle(o),
                ).grid(row=row, column=col, sticky="w", padx=(0, fs(4)))
            # If the last grid row has an empty column, tuck the custom
            # checkbox in there instead of always starting a fresh row -
            # saves a row of height whenever the preset count doesn't land
            # on an exact multiple of EXTRA_GRID_COLUMNS.
            n = len(options)
            last_row = (n - 1) // EXTRA_GRID_COLUMNS if n else 0
            last_row_count = n - last_row * EXTRA_GRID_COLUMNS
            if n and last_row_count < EXTRA_GRID_COLUMNS:
                custom_row, custom_col, custom_colspan = last_row, last_row_count, 1
            else:
                custom_row, custom_col, custom_colspan = last_row + (1 if n else 0), 0, EXTRA_GRID_COLUMNS
            tk.Checkbutton(
                extra_section, text="อื่นๆ (พิมพ์เอง):", variable=custom_var, font=("Tahoma", fs(9)),
                command=lambda: on_extra_toggle("__custom__"),
            ).grid(row=custom_row, column=custom_col, columnspan=custom_colspan, sticky="w", pady=(fs(4), 0))
            tk.Entry(
                extra_section, textvariable=custom_text_var, font=("Tahoma", fs(9)),
                validate="key", validatecommand=_custom_len_vcmd,
            ).grid(row=custom_row + 1, column=0, columnspan=EXTRA_GRID_COLUMNS, sticky="ew", padx=(fs(18), 0))

        def on_mode_change(*args):
            mode_var.set(mode_label_to_key[mode_display_var.get()])
            render_dose_fields()
            render_extra_fields()

        mode_display_var.trace_add("write", on_mode_change)
        render_dose_fields()
        render_extra_fields()

        def collect_into_d():
            if not drug1_var.get().strip():
                messagebox.showwarning("แจ้งเตือน", "กรุณาใส่ชื่อยา", parent=win)
                return False
            d["drug1"] = drug1_var.get().strip()
            d["barcode"] = barcode_var.get().strip()
            d["drug2"] = drug2_var.get().strip()
            new_note = note_var.get().strip()
            if new_note != d.get("note", ""):
                # Indication text changed - any Grok translation cached
                # against the old text is now stale, force a re-translate
                # next time this drug is previewed/printed on a
                # non-Thai label.
                d["note_en"] = ""
                d["note_mm"] = ""
            d["note"] = new_note
            d["qty"] = qty_var.get().strip()
            d["unit"] = unit_var.get().strip()
            d["per_day"] = per_day_var.get().strip()
            d["every_hr"] = every_hr_var.get().strip() if every_hr_enabled_var.get() else ""
            d["meal"] = meal_var.get().strip()
            d["times"] = [t for t, v in time_vars.items() if v.get()]
            d["extra_labels"] = collect_extra_labels()
            d["usage_mode"] = mode_var.get()
            # unit/meal are only meaningful for the field/mode combos that
            # actually show them - normalize the rest so stored data doesn't
            # carry stale values from a mode the pharmacist switched away from.
            if d["usage_mode"] == "topical":
                d["unit"] = ""
                d["meal"] = ""
            elif d["usage_mode"] == "drops":
                d["unit"] = "หยด"
                d["meal"] = ""
            d["status"] = "edited"
            return True

        def on_save():
            if not collect_into_d():
                return
            self.refresh_selected_list()
            win.destroy()

        def on_cancel():
            win.destroy()
            if is_new:
                # brand-new blank drug never got saved - drop it instead of
                # leaving an empty placeholder row in the print queue
                self.remove_drug(index)

        def on_save_to_db():
            if not collect_into_d():
                return
            if not messagebox.askyesno(
                "ยืนยัน",
                f"บันทึกวิธีกินของ '{d['drug1']}' ไว้ในเครื่องนี้ด้วยไหม?\n"
                "ครั้งต่อไปที่ค้นหาจะขึ้นเขียวอัตโนมัติ",
                parent=win,
            ):
                return
            save_db_btn.config(state="disabled")

            def worker():
                try:
                    new_id = save_product_med_info(d["idproduct"], d)
                    d["idproduct"] = new_id
                    # a row can be saved with only drug1 filled in (e.g. the
                    # pharmacist just typed a name and hit save without
                    # entering dosing info yet) - that's not "complete" any
                    # more than a bulk Excel import row is, so it must stay
                    # red, not turn green just because a DB write succeeded.
                    d["status"] = "db" if has_dosing_data(d) else "missing"
                    self.root.after(0, self.refresh_selected_list)
                    self.root.after(0, win.destroy)
                    self.root.after(0, lambda: self.status_var.set(
                        f"บันทึก {d['drug1']} แล้ว (ครั้งต่อไปจะขึ้นเขียวอัตโนมัติ)"))
                except Exception as e:
                    self.root.after(0, lambda: messagebox.showerror(
                        "ผิดพลาด", f"บันทึกไม่สำเร็จ: {e}", parent=win))
                    self.root.after(0, lambda: save_db_btn.config(state="normal"))

            threading.Thread(target=worker, daemon=True).start()

        btn_frame = tk.Frame(win)
        btn_frame.pack(pady=fs(8))
        save_db_btn = tk.Button(btn_frame, text="\U0001F4BE บันทึกในเครื่อง",
                                 font=("Tahoma", fs(10), "bold"), bg="#1a5a9a", fg="white",
                                 command=on_save_to_db)
        save_db_btn.pack(side="left", padx=fs(6))
        tk.Button(btn_frame, text="บันทึกครั้งนี้", font=("Tahoma", fs(10), "bold"),
                  bg="#c07a17", fg="white", command=on_save).pack(side="left", padx=fs(6))
        tk.Button(btn_frame, text="ยกเลิก", font=("Tahoma", fs(11)), command=on_cancel).pack(side="left", padx=fs(6))

        win.protocol("WM_DELETE_WINDOW", on_cancel)
        win.lift()
        win.focus_force()

    # ---------------------------------------------------------------- mobile queue

    def _copy_queue_url(self, copied_var):
        self.root.clipboard_clear()
        self.root.clipboard_append(self.queue_url)
        copied_var.set(f"คัดลอก {self.queue_url} แล้ว")

    def open_queue_dialog(self):
        win = tk.Toplevel(self.root)
        win.title("คิวจากมือถือ")
        win.geometry(f"{fs(440)}x{fs(420)}")
        win.transient(self.root)
        win.grab_set()

        tk.Label(
            win, text="ให้พนักงานเข้า URL ด้านล่างผ่าน WiFi วงเดียวกับเครื่องนี้",
            font=("Tahoma", fs(9), "bold"), fg="#1a5a9a", wraplength=fs(400), justify="left",
        ).pack(anchor="w", padx=fs(10), pady=(fs(10), fs(2)))

        url_row = tk.Frame(win)
        url_row.pack(fill="x", padx=fs(10), pady=(0, fs(2)))
        url_var = tk.StringVar(value=self.queue_url or "ไม่สามารถเปิด local server ได้ (พอร์ตไม่ว่างเลย)")
        url_entry = tk.Entry(
            url_row, textvariable=url_var, font=("Tahoma", fs(11), "bold"), fg="#1a5a9a",
            state="readonly", readonlybackground="white", relief="solid", bd=1,
        )
        url_entry.pack(side="left", fill="x", expand=True, ipady=fs(3))
        if self.queue_url:
            tk.Button(
                url_row, text="📋 คัดลอก", font=("Tahoma", fs(9)),
                command=lambda: self._copy_queue_url(copied_var),
            ).pack(side="left", padx=(fs(4), 0))

        copied_var = tk.StringVar(value="")
        tk.Label(win, textvariable=copied_var, font=("Tahoma", fs(8)), fg="#0a7a2f").pack(
            anchor="w", padx=fs(10), pady=(0, fs(2))
        )
        tk.Label(
            win, text="รายการที่พนักงานส่งมาจากมือถือ ดับเบิลคลิกเพื่อดึงมาพิมพ์",
            font=("Tahoma", fs(9)), fg="#555", wraplength=fs(400), justify="left",
        ).pack(anchor="w", padx=fs(10), pady=(0, fs(4)))

        listbox = tk.Listbox(win, font=("Tahoma", fs(10)))
        listbox.pack(fill="both", expand=True, padx=fs(10), pady=fs(4))

        status_var = tk.StringVar(value="")
        tk.Label(win, textvariable=status_var, font=("Tahoma", fs(9)), fg="#a00",
                 wraplength=fs(400), justify="left").pack(padx=fs(10), pady=(0, fs(6)))

        jobs = []

        def refresh():
            listbox.delete(0, tk.END)
            try:
                jobs_raw = storage.list_queue_jobs()
            except Exception as e:
                status_var.set(f"โหลดคิวไม่สำเร็จ: {e}")
                return
            jobs.clear()
            jobs.extend(jobs_raw)
            for idx, job in enumerate(jobs):
                # a job is "complete" (green) only if every one of its drugs
                # still has a saved template locally right now - dosing info
                # may have changed (or been deleted) since it was submitted
                ids = [d.get("idproduct") for d in job["drugs"] if d.get("idproduct")]
                complete = bool(ids)
                for i in ids:
                    info = storage.get_template(i)
                    if not has_dosing_data(info):
                        complete = False
                        break
                when = job["submitted_at"].split("T")[-1][:8] if "T" in job["submitted_at"] else job["submitted_at"]
                # patient_name doubles as the submitting staff's name here -
                # the mobile page's staff picker writes into that field since
                # there's no separate patient-name collection anymore
                who = job["patient_name"] or "(ไม่ระบุชื่อ)"
                listbox.insert(tk.END, f"{who} ({len(job['drugs'])} รายการ) - {when}")
                listbox.itemconfig(idx, fg="#0a7a2f" if complete else "#c0392b")
            status_var.set(f"พบ {len(jobs)} รายการ" if jobs else "ยังไม่มีรายการส่งเข้ามา")

        def on_pick(event=None):
            sel = listbox.curselection()
            if not sel:
                return
            job = jobs[sel[0]]
            claimed = storage.claim_queue_job(job["id"])
            if not claimed:
                status_var.set("รายการนี้ถูกดึงไปพิมพ์แล้ว")
                refresh()
                return

            # Never fully trust the drug snapshot from the moment it was
            # submitted - re-resolve each drug by idproduct against the local
            # SQLite templates right now, same principle as a normal local
            # pick (add_drug). Dosing info may have changed since submission.
            resolved = []
            for d in claimed["drugs"]:
                entry = dict(d)
                entry.setdefault("print_qty", 1)
                idproduct = entry.get("idproduct")
                info = storage.get_template(idproduct) if idproduct else None
                if has_dosing_data(info):
                    entry.update(info)
                    entry["status"] = "db"
                else:
                    entry.setdefault("status", "missing")
                    entry.setdefault("usage_mode", "oral")
                resolved.append(entry)

            self.selected_drugs.extend(resolved)
            self.refresh_selected_list()
            if claimed["patient_name"]:
                self._queue_patient_name = claimed["patient_name"]
            if claimed.get("has_allergy"):
                self.allergy_var.set(True)
            win.destroy()
            self.status_var.set(f"ดึงคิวมาแล้ว ({len(resolved)} รายการยา) - กด ยืนยันรายการยา เพื่อพิมพ์")

        listbox.bind("<Double-Button-1>", on_pick)
        tk.Button(win, text="🔄 รีเฟรช", font=("Tahoma", fs(9)), command=refresh).pack(pady=(0, fs(8)))

        win.lift()
        win.focus_force()
        refresh()

    # ---------------------------------------------------------------- print history

    def open_print_history_dialog(self, pick_mode=False):
        win = tk.Toplevel(self.root)
        win.title("แฟ้มประวัติการจ่ายยา - เลือกชื่อลูกค้า" if pick_mode else "แฟ้มประวัติการจ่ายยา")
        win.geometry(f"{fs(480)}x{fs(600)}")
        win.transient(self.root)
        win.grab_set()

        tk.Label(
            win, text="เก็บถาวรทุกรายการ (ไม่ลบทิ้งอัตโนมัติ) - ค้นหาชื่อ/เบอร์โทรเพื่อดูประวัติย้อนหลังทั้งหมด "
                       "ดับเบิลคลิกที่ยาแต่ละตัวเพื่อดูวิธีกิน หรือกด \"ใช้ซ้ำ\" เพื่อโหลดกลับมาพิมพ์ใหม่",
            font=("Tahoma", fs(9)), fg="#555", wraplength=fs(440), justify="left",
        ).pack(anchor="w", padx=fs(10), pady=(fs(10), fs(6)))

        tk.Label(
            win, text="ค้นหาชื่อ/เบอร์โทร (ว่างไว้ = แสดง 24 ชม.ล่าสุด)", font=("Tahoma", fs(9)), fg="#555",
        ).pack(anchor="w", padx=fs(10))
        search_row = tk.Frame(win)
        search_row.pack(fill="x", padx=fs(10), pady=(0, fs(6)))
        search_var = tk.StringVar()
        tk.Entry(search_row, textvariable=search_var, font=("Tahoma", fs(10))).pack(side="left", fill="x", expand=True)
        tk.Button(
            search_row, text="✕", font=("Tahoma", fs(9), "bold"), fg="white", bg="#555555", width=2,
            command=lambda: search_var.set(""),
        ).pack(side="left", padx=(fs(4), 0))

        tk.Label(win, text="รายการ (ชื่อลูกค้า - เบอร์โทร - เวลา - จำนวนยา)", font=("Tahoma", fs(10), "bold")).pack(anchor="w", padx=fs(10))
        # exportselection=False on BOTH listboxes - without this, Tkinter
        # treats all listboxes in the app as sharing one "selection", so
        # selecting a row in drug_list silently clears job_list's selection,
        # which fires job_list's <<ListboxSelect>> again and wipes drug_list
        # right back out from under the click (looked like items vanishing
        # and double-click never registering - this was a real bug, not a
        # user-error report).
        job_list = tk.Listbox(win, font=("Tahoma", fs(9)), height=8, exportselection=False)
        job_list.pack(fill="x", padx=fs(10), pady=(fs(2), fs(4)))

        job_btn_row = tk.Frame(win)
        job_btn_row.pack(fill="x", padx=fs(10), pady=(0, fs(8)))
        tk.Button(
            job_btn_row, text="👁 ซ่อน/แสดง (อ่านแล้ว)", font=("Tahoma", fs(8)),
            command=lambda: toggle_hidden(),
        ).pack(side="left")
        hide_all_btn = tk.Button(
            job_btn_row, text="🗑 ซ่อนทั้งหมด (ไม่ลบถาวร)", font=("Tahoma", fs(8), "bold"), fg="white", bg="#b03a2e",
            command=lambda: toggle_hide_all(),
        )
        hide_all_btn.pack(side="left", padx=(fs(6), 0))
        tk.Button(
            job_btn_row, text="✕ ลบถาวร", font=("Tahoma", fs(8), "bold"), fg="white", bg="#b03a2e",
            command=lambda: delete_job(),
        ).pack(side="left", padx=(fs(6), 0))

        tk.Label(win, text="ยาในรายการนี้ (ดับเบิลคลิกดูวิธีกิน)", font=("Tahoma", fs(10), "bold")).pack(anchor="w", padx=fs(10))
        drug_list = tk.Listbox(win, font=("Tahoma", fs(9)), exportselection=False)
        drug_list.pack(fill="both", expand=True, padx=fs(10), pady=(fs(2), fs(6)))

        jobs = []

        def format_when(printed_at):
            if printed_at is None:
                return "(ยังไม่เคยพิมพ์)"
            try:
                dt = datetime.fromisoformat(printed_at)
                return dt.strftime("%d/%m/%Y %H:%M")
            except Exception:
                return printed_at

        def refresh():
            term = search_var.get().strip()
            try:
                if term:
                    jobs_raw = storage.search_print_jobs(term)
                else:
                    # Archived jobs (bulk "start a new day" action) are kept
                    # out of the default (unsearched) view so the button
                    # actually clears the day's list - but they still surface
                    # via a search (term branch above doesn't filter
                    # archived). hidden (per-row "mark as read") is a
                    # SEPARATE flag - never filtered, only dims the row below.
                    jobs_raw = [j for j in storage.list_print_jobs(hours=24) if not j["archived"]]
            except Exception as e:
                messagebox.showerror("โหลดไม่สำเร็จ", str(e), parent=win)
                return
            if pick_mode and term:
                # A patient profile (with allergy history) can be created
                # directly in "ประวัติผู้ป่วย" without ever having a label
                # printed for them - without this merge they'd never show up
                # here since this list is normally sourced from print job
                # history, making them impossible to pick as the current
                # customer and impossible to auto-detect their allergy.
                existing_ids = {j.get("patient_id") for j in jobs_raw if j.get("patient_id")}
                try:
                    patients = storage.search_patients(term)
                except Exception:
                    patients = []
                for p in patients:
                    if p["id"] in existing_ids:
                        continue
                    jobs_raw.append({
                        "patient_name": p["name"], "customer_phone": p.get("phone", ""),
                        "printed_at": None, "drugs": [], "patient_id": p["id"],
                        "hidden": False, "archived": False,
                    })
            jobs.clear()
            jobs.extend(jobs_raw)
            job_list.delete(0, tk.END)
            for idx, job in enumerate(jobs):
                when = format_when(job["printed_at"])
                who = job["patient_name"] or "(ไม่ระบุชื่อ)"
                phone = f" - {job['customer_phone']}" if job.get("customer_phone") else ""
                job_list.insert(tk.END, f"{who}{phone} - {when} - {len(job['drugs'])} รายการ")
                if job["hidden"]:
                    job_list.itemconfig(idx, fg="#aaaaaa")
            drug_list.delete(0, tk.END)

        def on_job_select(event=None):
            sel = job_list.curselection()
            drug_list.delete(0, tk.END)
            if not sel:
                return
            job = jobs[sel[0]]
            for d in job["drugs"]:
                label = f"{d['drug1']} ({d['drug2']})" if d.get("drug2") else d["drug1"]
                drug_list.insert(tk.END, f"{label}  x{d.get('print_qty', 1)}")

        def on_drug_pick(event=None):
            job_sel = job_list.curselection()
            drug_sel = drug_list.curselection()
            if not job_sel or not drug_sel:
                return
            job = jobs[job_sel[0]]
            d = job["drugs"][drug_sel[0]]
            dose_text, line2 = compute_dose_lines(d)
            lines = []
            if d.get("note"):
                lines.append(d["note"])
            lines.append(dose_text)
            if line2:
                lines.append(line2)
            if d.get("extra_labels"):
                lines.append(" ".join(f"**{e}**" for e in d["extra_labels"]))
            show_dose_popup(d["drug1"], "\n".join(lines))

        def show_dose_popup(title, text):
            # messagebox.showinfo can't have a custom font - use a plain
            # Toplevel instead so the dosing text can be shown 2x normal
            # size (easier to read at a glance while counting out pills).
            popup = tk.Toplevel(win)
            popup.title(title)
            popup.transient(win)
            popup.grab_set()
            tk.Label(
                popup, text=title, font=("Tahoma", fs(14), "bold"), wraplength=fs(380), justify="left",
            ).pack(padx=fs(20), pady=(fs(16), fs(4)))
            tk.Label(
                popup, text=text, font=("Tahoma", fs(20)), wraplength=fs(380), justify="left",
            ).pack(padx=fs(20), pady=(fs(4), fs(16)))
            tk.Button(popup, text="ปิด", font=("Tahoma", fs(11)), command=popup.destroy).pack(pady=(0, fs(14)))
            popup.lift()
            popup.focus_force()

        def reuse_job():
            sel = job_list.curselection()
            if not sel:
                messagebox.showwarning("แจ้งเตือน", "กรุณาเลือกรายการก่อน", parent=win)
                return
            job = jobs[sel[0]]
            # never fully trust the snapshot from when it was printed -
            # re-check against the current local templates, same principle
            # as loading a Favorite or claiming a mobile-queue job
            for d in job["drugs"]:
                entry = dict(d)
                idproduct = entry.get("idproduct")
                info = storage.get_template(idproduct) if idproduct else None
                if has_dosing_data(info):
                    entry.update(info)
                    entry["status"] = "db"
                else:
                    entry.setdefault("status", "missing")
                self.selected_drugs.append(entry)
            self.refresh_selected_list()
            win.destroy()
            self.status_var.set(f"โหลดรายการซ้ำแล้ว ({len(job['drugs'])} รายการยา)")

        def show_patient_alert_popup(patient_id):
            """Surfaces allergy info the moment a returning customer is
            picked - the whole point of linking to a patient profile is
            useless if the pharmacist still has to go dig for it manually
            in a separate dialog before dispensing."""
            try:
                patient = storage.get_patient(patient_id)
            except Exception:
                patient = None
            if not patient:
                return
            popup = tk.Toplevel(win)
            popup.title("ประวัติผู้ป่วย - ตรวจสอบก่อนจ่ายยา")
            popup.transient(win)
            popup.grab_set()

            hn_part = f"  [HN {patient['hn_code']}]" if patient.get("hn_code") else ""
            tk.Label(
                popup, text=patient["name"] + hn_part, font=("Tahoma", fs(13), "bold"),
            ).pack(anchor="w", padx=fs(14), pady=(fs(12), fs(4)))

            tk.Label(popup, text="ประวัติแพ้ยา", font=("Tahoma", fs(10), "bold")).pack(anchor="w", padx=fs(14))
            allergy_note = (patient.get("allergy_note") or "").strip()
            if allergy_note:
                tk.Label(
                    popup, text=allergy_note, font=("Tahoma", fs(11), "bold"), fg="#b03a2e",
                    wraplength=fs(360), justify="left",
                ).pack(anchor="w", padx=fs(14), pady=(fs(2), fs(10)))
            else:
                tk.Label(
                    popup, text="ไม่มีประวัติแพ้ยาบันทึกไว้", font=("Tahoma", fs(10)), fg="#666",
                ).pack(anchor="w", padx=fs(14), pady=(fs(2), fs(10)))

            def open_full_profile():
                popup.destroy()
                win.grab_release()  # avoid competing with the new dialog's own grab
                self.open_patient_profile_dialog(preload_patient_id=patient_id)

            btn_row = tk.Frame(popup)
            btn_row.pack(pady=(0, fs(12)))
            tk.Button(
                btn_row, text="📁 เปิดประวัติผู้ป่วยเต็ม", font=("Tahoma", fs(9), "bold"),
                bg="#1a5a9a", fg="white", command=open_full_profile,
            ).pack(side="left", padx=fs(4))
            tk.Button(btn_row, text="ปิด", font=("Tahoma", fs(9)), command=popup.destroy).pack(side="left", padx=fs(4))

            popup.lift()
            popup.focus_force()

        def pick_name():
            sel = job_list.curselection()
            if not sel:
                messagebox.showwarning("แจ้งเตือน", "กรุณาเลือกรายการก่อน", parent=win)
                return
            job = jobs[sel[0]]
            name = job["patient_name"] or ""
            phone = job.get("customer_phone") or ""
            self._queue_patient_name = name
            self._queue_patient_phone = phone
            # Re-resolve rather than trust job["patient_id"] as-is: that job
            # row may predate this column, or the patient may have been
            # saved to a profile only after this particular print happened.
            # Still never auto-creates - stays None if there's no
            # unambiguous existing profile for this exact name+phone.
            self._queue_patient_id = job.get("patient_id") or storage.find_patient_id(name, phone)
            self.selected_customer_var.set("ลค: " + name + (f" ({phone})" if phone else ""))
            self.status_var.set(f"เลือกชื่อลูกค้า '{name}' แล้ว - จะใส่ให้อัตโนมัติตอนกดยืนยันพิมพ์")
            if self._queue_patient_id:
                # Auto-tick "แพ้ยา" the moment we know this patient has any
                # allergy history at all, regardless of how many drugs are in
                # it - the pharmacist shouldn't have to remember to tick it
                # manually after already picking a known allergic patient.
                try:
                    patient = storage.get_patient(self._queue_patient_id)
                except Exception:
                    patient = None
                if patient and (patient.get("allergy_note") or "").strip():
                    self.allergy_var.set(True)
                show_patient_alert_popup(self._queue_patient_id)

        def toggle_hidden():
            sel = job_list.curselection()
            if not sel:
                messagebox.showwarning("แจ้งเตือน", "กรุณาเลือกรายการก่อน", parent=win)
                return
            job = jobs[sel[0]]
            storage.set_print_job_hidden(job["id"], not job["hidden"])
            refresh()

        def delete_job():
            sel = job_list.curselection()
            if not sel:
                messagebox.showwarning("แจ้งเตือน", "กรุณาเลือกรายการก่อน", parent=win)
                return
            job = jobs[sel[0]]
            who = job["patient_name"] or "(ไม่ระบุชื่อ)"
            if not messagebox.askyesno("ยืนยัน", f"ลบรายการของ '{who}' ถาวรใช่ไหม? กู้คืนไม่ได้", parent=win):
                return
            storage.delete_print_job(job["id"])
            refresh()

        all_archived_state = {"archived": False}

        def toggle_hide_all():
            new_state = not all_archived_state["archived"]
            if new_state:
                if not messagebox.askyesno(
                    "ยืนยัน",
                    "ซ่อนรายการทั้งหมดที่เห็นตอนนี้ใช่ไหม? (ไม่ลบถาวร - ยังค้นหาย้อนหลังได้ปกติ "
                    "กดปุ่มนี้ซ้ำเพื่อแสดงกลับได้ทุกเมื่อ)",
                    parent=win,
                ):
                    return
            storage.set_all_print_jobs_archived(new_state)
            all_archived_state["archived"] = new_state
            hide_all_btn.config(text="👁 แสดงทั้งหมด" if new_state else "🗑 ซ่อนทั้งหมด (ไม่ลบถาวร)")
            refresh()

        job_list.bind("<<ListboxSelect>>", on_job_select)
        drug_list.bind("<Double-Button-1>", on_drug_pick)
        search_var.trace_add("write", lambda *a: refresh())

        btn_row = tk.Frame(win)
        btn_row.pack(pady=(0, fs(8)))
        if pick_mode:
            tk.Button(
                btn_row, text="✓ เลือกชื่อนี้", font=("Tahoma", fs(9), "bold"),
                bg="#1a5a9a", fg="white", command=pick_name,
            ).pack(side="left", padx=fs(4))
        tk.Button(
            btn_row, text="🔁 ใช้ซ้ำ (โหลดกลับมาพิมพ์)", font=("Tahoma", fs(9), "bold"),
            bg="#1a7a4a", fg="white", command=reuse_job,
        ).pack(side="left", padx=fs(4))
        tk.Button(btn_row, text="🔄 รีเฟรช", font=("Tahoma", fs(9)), command=refresh).pack(side="left", padx=fs(4))

        win.lift()
        win.focus_force()
        refresh()

    # ---------------------------------------------------------------- patient profile

    def open_patient_profile_dialog(self, preload_patient_id=None):
        win = tk.Toplevel(self.root)
        win.title("ประวัติผู้ป่วย")
        win_w = fs(520)
        win_h = min(fs(660), win.winfo_screenheight() - fs(80))
        win_x = (win.winfo_screenwidth() - win_w) // 2
        win_y = max(0, (win.winfo_screenheight() - win_h) // 2 - fs(20))
        win.geometry(f"{win_w}x{win_h}+{win_x}+{win_y}")
        win.transient(self.root)
        win.grab_set()

        tk.Label(win, text="ค้นหาชื่อ/เบอร์โทร", font=("Tahoma", fs(10), "bold")).pack(anchor="w", padx=fs(10), pady=(fs(10), fs(2)))
        search_row = tk.Frame(win)
        search_row.pack(fill="x", padx=fs(10), pady=(0, fs(6)))
        search_var = tk.StringVar()
        search_entry = tk.Entry(search_row, textvariable=search_var, font=("Tahoma", fs(10)))
        search_entry.pack(side="left", fill="x", expand=True)
        tk.Button(
            search_row, text="✕", font=("Tahoma", fs(9), "bold"), fg="white", bg="#555555", width=2,
            command=lambda: (search_var.set(""), result_list.delete(0, tk.END)),
        ).pack(side="left", padx=(fs(4), 0))
        tk.Button(search_row, text="ค้นหา", font=("Tahoma", fs(9)), command=lambda: do_search()).pack(side="left", padx=(fs(4), 0))
        tk.Button(
            search_row, text="📋 HN ทั้งหมด", font=("Tahoma", fs(9)),
            command=lambda: self.open_all_hn_dialog(win, load_patient),
        ).pack(side="left", padx=(fs(4), 0))

        result_list = tk.Listbox(win, font=("Tahoma", fs(9)), height=4, exportselection=False)
        result_list.pack(fill="x", padx=fs(10), pady=(0, fs(8)))

        divider = tk.Frame(win, height=2, bg="#ccc")
        divider.pack(fill="x", padx=fs(10), pady=(0, fs(6)))

        patient_name_var = tk.StringVar(value="(ยังไม่ได้เลือกผู้ป่วย)")
        tk.Label(win, textvariable=patient_name_var, font=("Tahoma", fs(13), "bold"), fg="#1a7a4a").pack(
            anchor="w", padx=fs(10), pady=(0, fs(4))
        )

        tk.Label(win, text="ประวัติแพ้ยา", font=("Tahoma", fs(10), "bold")).pack(anchor="w", padx=fs(10))
        tk.Label(
            win, text="(ถ้ามีหลายตัว ให้คั่นด้วย comma เช่น Amoxy, Diclofenac)",
            font=("Tahoma", fs(8)), fg="#777",
        ).pack(anchor="w", padx=fs(10))
        allergy_text = tk.Text(win, font=("Tahoma", fs(10)), height=3, bd=1, relief="solid")
        allergy_text.pack(fill="x", padx=fs(10), pady=(fs(2), fs(4)))
        tk.Button(
            win, text="💾 บันทึกประวัติแพ้ยา", font=("Tahoma", fs(9), "bold"), bg="#1a7a4a", fg="white",
            command=lambda: save_allergy(),
        ).pack(anchor="w", padx=fs(10), pady=(0, fs(8)))

        tk.Label(win, text="ประวัติการซื้อทั้งหมด", font=("Tahoma", fs(10), "bold")).pack(anchor="w", padx=fs(10))
        purchase_list = tk.Listbox(win, font=("Tahoma", fs(9)), height=4, exportselection=False)
        purchase_list.pack(fill="x", padx=fs(10), pady=(fs(2), fs(8)))

        doc_header_row = tk.Frame(win)
        doc_header_row.pack(fill="x", padx=fs(10))
        tk.Label(doc_header_row, text="เอกสารประกอบ (รูป+หมายเหตุ)", font=("Tahoma", fs(10), "bold")).pack(side="left")
        tk.Button(
            doc_header_row, text="📷 เพิ่มรูป+หมายเหตุ", font=("Tahoma", fs(8), "bold"), bg="#1a5a9a", fg="white",
            command=lambda: upload_doc(),
        ).pack(side="right")

        doc_list_outer = tk.Frame(win, bd=1, relief="solid")
        doc_list_outer.pack(fill="both", expand=True, padx=fs(10), pady=(fs(2), fs(8)))
        doc_canvas = tk.Canvas(doc_list_outer, highlightthickness=0)
        doc_scroll = tk.Scrollbar(doc_list_outer, orient="vertical", command=doc_canvas.yview)
        doc_rows_frame = tk.Frame(doc_canvas)
        doc_rows_frame.bind(
            "<Configure>", lambda e: doc_canvas.configure(scrollregion=doc_canvas.bbox("all"))
        )
        doc_canvas_window = doc_canvas.create_window((0, 0), window=doc_rows_frame, anchor="nw")
        doc_canvas.bind(
            "<Configure>", lambda e: doc_canvas.itemconfig(doc_canvas_window, width=e.width)
        )
        doc_canvas.configure(yscrollcommand=doc_scroll.set)
        doc_canvas.pack(side="left", fill="both", expand=True)
        doc_scroll.pack(side="right", fill="y")

        current_patient = {"id": None, "name": "", "phone": ""}
        results = []
        purchase_jobs = []
        docs = []

        def format_when(iso_str):
            try:
                return datetime.fromisoformat(iso_str).strftime("%d/%m/%Y %H:%M")
            except Exception:
                return iso_str

        def load_patient(patient_id):
            p = storage.get_patient(patient_id)
            if not p:
                return
            current_patient["id"] = p["id"]
            current_patient["name"] = p["name"]
            current_patient["phone"] = p["phone"]
            hn_part = f"  [HN {p['hn_code']}]" if p.get("hn_code") else ""
            patient_name_var.set(p["name"] + (f" ({p['phone']})" if p["phone"] else "") + hn_part)
            allergy_text.delete("1.0", "end")
            allergy_text.insert("1.0", p["allergy_note"])
            refresh_purchase_history()
            refresh_documents()

        def do_search():
            term = search_var.get().strip()
            if not term:
                return
            results.clear()
            results.extend(storage.search_patients(term))
            result_list.delete(0, tk.END)
            for p in results:
                phone_part = f" - {p['phone']}" if p["phone"] else ""
                result_list.insert(tk.END, f"{p['name']}{phone_part}")
            if not results:
                if messagebox.askyesno(
                    "ไม่พบผู้ป่วย", f"ไม่พบผู้ป่วยชื่อ/เบอร์ '{term}' ต้องการสร้างประวัติใหม่ไหม?", parent=win,
                ):
                    # crude heuristic: all-digit-ish term -> phone, else name
                    if term.replace("-", "").replace(" ", "").isdigit():
                        pid = storage.find_or_create_patient("", term)
                    else:
                        pid = storage.find_or_create_patient(term, "")
                    load_patient(pid)

        def on_result_select(event=None):
            sel = result_list.curselection()
            if not sel:
                return
            load_patient(results[sel[0]]["id"])

        def refresh_purchase_history():
            purchase_list.delete(0, tk.END)
            purchase_jobs.clear()
            if not current_patient["id"]:
                return
            jobs = storage.list_print_jobs_for_patient(
                current_patient["name"], current_patient["phone"], patient_id=current_patient["id"]
            )
            purchase_jobs.extend(jobs)
            if not jobs:
                purchase_list.insert(tk.END, "(ยังไม่มีประวัติการซื้อ)")
                return
            for job in jobs:
                when = format_when(job["printed_at"])
                drug_names = ", ".join(d["drug1"] for d in job["drugs"][:3])
                if len(job["drugs"]) > 3:
                    drug_names += ", ..."
                purchase_list.insert(tk.END, f"{when} - {drug_names}")

        def save_allergy():
            if not current_patient["id"]:
                messagebox.showwarning("แจ้งเตือน", "กรุณาค้นหา/เลือกผู้ป่วยก่อน", parent=win)
                return
            note = allergy_text.get("1.0", "end").strip()
            storage.update_patient_allergy(current_patient["id"], note)
            self.status_var.set(f"บันทึกประวัติแพ้ยาของ {current_patient['name']} แล้ว")

        def refresh_documents():
            for child in doc_rows_frame.winfo_children():
                child.destroy()
            docs.clear()
            if not current_patient["id"]:
                return
            docs.extend(storage.list_patient_documents(current_patient["id"]))
            if not docs:
                tk.Label(doc_rows_frame, text="(ยังไม่มีเอกสาร)", font=("Tahoma", fs(9))).pack(
                    anchor="w", padx=fs(4), pady=fs(4)
                )
                return
            for d in docs:
                when = format_when(d["uploaded_at"])
                note_part = f" - {d['note']}" if d["note"] else ""
                row = tk.Frame(doc_rows_frame)
                row.pack(fill="x", padx=fs(4), pady=fs(2))
                tk.Label(
                    row, text=f"{when}{note_part}", font=("Tahoma", fs(9)),
                    anchor="w", justify="left", wraplength=fs(320),
                ).pack(side="left", fill="x", expand=True)
                tk.Button(
                    row, text="✕", font=("Tahoma", fs(7), "bold"), fg="white", bg="#b03a2e", width=2,
                    command=lambda d=d: delete_doc(d),
                ).pack(side="right", padx=(fs(3), 0))
                tk.Button(
                    row, text="👁", font=("Tahoma", fs(7)), width=2,
                    command=lambda d=d: view_doc(d),
                ).pack(side="right", padx=(fs(2), 0))

        def upload_doc():
            if not current_patient["id"]:
                messagebox.showwarning("แจ้งเตือน", "กรุณาค้นหา/เลือกผู้ป่วยก่อน", parent=win)
                return
            path = filedialog.askopenfilename(
                title="เลือกรูปภาพ", parent=win,
                filetypes=[("Image files", "*.jpg *.jpeg *.png *.bmp *.gif")],
            )
            if not path:
                return
            note = ask_upload_note(win)
            if note is None:
                return
            try:
                with open(path, "rb") as f:
                    image_bytes = f.read()
                storage.add_patient_document(current_patient["id"], image_bytes, note)
            except Exception as e:
                messagebox.showerror("อัปโหลดไม่สำเร็จ", str(e), parent=win)
                return
            refresh_documents()

        def view_doc(d):
            try:
                os.startfile(d["full_path"])
            except Exception as e:
                messagebox.showerror("เปิดไม่สำเร็จ", str(e), parent=win)

        def delete_doc(d):
            if not messagebox.askyesno("ยืนยัน", "ลบเอกสารนี้ถาวรใช่ไหม?", parent=win):
                return
            storage.delete_patient_document(d["id"])
            refresh_documents()

        result_list.bind("<<ListboxSelect>>", on_result_select)
        search_entry.bind("<Return>", lambda e: do_search())

        if preload_patient_id:
            load_patient(preload_patient_id)

        win.lift()
        win.focus_force()
        search_entry.focus_set()

    def open_all_hn_dialog(self, parent_win, on_pick):
        """Management list of every patient/HN record - separate from the
        search box above (which is for finding ONE patient to work with).
        This is for browsing/cleaning up the whole list: sort by name or by
        HN, delete a mistaken/duplicate entry, or wipe everything at once
        for a store that just started using patient profiles and wants a
        clean slate (test entries, wrong HN numbering, etc.)."""
        win = tk.Toplevel(parent_win)
        win.title("รายชื่อ HN ทั้งหมด")
        win.geometry(f"{fs(420)}x{fs(520)}")
        win.transient(parent_win)
        win.grab_set()

        tk.Label(
            win, text="รูปแบบ HN (ปี/เดือน/วัน เลขรันจาก 00001) - เช่น 260715-00001 คือวันที่ 15 ก.ค. 2026 (ค.ศ.) "
                       "คนที่ 1 ของวันนั้น เลขรันจะเริ่มนับ 00001 ใหม่ทุกวัน",
            font=("Tahoma", fs(8)), fg="#666", wraplength=fs(400), justify="left",
        ).pack(anchor="w", padx=fs(10), pady=(0, fs(2)))

        top_row = tk.Frame(win)
        top_row.pack(fill="x", padx=fs(10), pady=(fs(10), fs(4)))
        tk.Label(top_row, text="เรียงตาม:", font=("Tahoma", fs(9))).pack(side="left")
        sort_var = tk.StringVar(value="name")
        tk.Radiobutton(
            top_row, text="ชื่อ", variable=sort_var, value="name", font=("Tahoma", fs(9)),
            command=lambda: refresh(),
        ).pack(side="left")
        tk.Radiobutton(
            top_row, text="HN", variable=sort_var, value="hn_code", font=("Tahoma", fs(9)),
            command=lambda: refresh(),
        ).pack(side="left")

        hn_list = tk.Listbox(win, font=("Tahoma", fs(9)), exportselection=False)
        hn_list.pack(fill="both", expand=True, padx=fs(10), pady=(fs(2), fs(6)))

        patients = []

        def refresh():
            patients.clear()
            patients.extend(storage.list_all_patients(order_by=sort_var.get()))
            hn_list.delete(0, tk.END)
            for p in patients:
                phone_part = f" - {p['phone']}" if p["phone"] else ""
                hn_code = p["hn_code"] or "(ไม่มี HN)"
                hn_list.insert(tk.END, f"{hn_code}  |  {p['name']}{phone_part}")

        def selected_patient():
            sel = hn_list.curselection()
            if not sel:
                messagebox.showwarning("แจ้งเตือน", "กรุณาเลือกรายการก่อน", parent=win)
                return None
            return patients[sel[0]]

        def on_pick_row(event=None):
            p = selected_patient()
            if not p:
                return
            on_pick(p["id"])
            win.destroy()

        def delete_selected():
            p = selected_patient()
            if not p:
                return
            hn_code = p["hn_code"] or "(ไม่มี HN)"
            if not messagebox.askyesno(
                "ยืนยัน",
                f"ลบ {hn_code} - {p['name']} ถาวรใช่ไหม?\n"
                "(ประวัติการพิมพ์ฉลากเดิมของคนนี้จะยังอยู่ แค่ไม่ผูกกับโปรไฟล์นี้แล้ว)",
                parent=win,
            ):
                return
            storage.delete_patient(p["id"])
            refresh()

        def delete_all():
            if not messagebox.askyesno(
                "ยืนยันการลบทั้งหมด",
                f"ลบข้อมูลผู้ป่วยทั้งหมด ({len(patients)} คน) รวม HN ทุกเลขถาวรใช่ไหม?\n"
                "เอกสารแนบทั้งหมดจะถูกลบด้วย - กู้คืนไม่ได้\n"
                "(ประวัติการพิมพ์ฉลากเดิมจะยังอยู่ แค่ไม่ผูกกับโปรไฟล์ผู้ป่วยคนไหนแล้ว "
                "เลข HN ใหม่จะเริ่มนับจาก 00001 อีกครั้ง)",
                parent=win,
            ):
                return
            if not messagebox.askyesno(
                "ยืนยันอีกครั้ง", "แน่ใจนะ? การลบนี้กู้คืนไม่ได้", parent=win,
            ):
                return
            storage.delete_all_patients()
            refresh()
            self.status_var.set("ลบข้อมูลผู้ป่วย/HN ทั้งหมดแล้ว")

        def assign_missing():
            # Self-serve fix for rows like "(ไม่มี HN)" - can happen if a
            # patient was created by an older build of this app (from
            # before the hn_code column existed) still running against the
            # same data.db, or any other gap. Safe to click any time: only
            # touches rows that are still NULL, never reassigns an existing
            # code.
            n = storage.backfill_patient_hn_codes()
            refresh()
            if n:
                messagebox.showinfo("เสร็จแล้ว", f"เติม HN ให้ผู้ป่วย {n} คนที่ยังไม่มีแล้ว", parent=win)
            else:
                messagebox.showinfo("แจ้งเตือน", "ทุกคนมี HN อยู่แล้ว ไม่มีอะไรต้องเติม", parent=win)

        hn_list.bind("<Double-Button-1>", on_pick_row)

        btn_row = tk.Frame(win)
        btn_row.pack(fill="x", padx=fs(10), pady=(0, fs(4)))
        tk.Button(
            btn_row, text="✓ เลือก", font=("Tahoma", fs(9), "bold"), bg="#1a5a9a", fg="white",
            command=on_pick_row,
        ).pack(side="left", padx=(0, fs(4)))
        tk.Button(
            btn_row, text="🗑 ลบ", font=("Tahoma", fs(9), "bold"), bg="#b03a2e", fg="white",
            command=delete_selected,
        ).pack(side="left", padx=(0, fs(4)))
        tk.Button(
            btn_row, text="🗑 ลบทั้งหมด", font=("Tahoma", fs(9), "bold"), bg="#b03a2e", fg="white",
            command=delete_all,
        ).pack(side="right")

        btn_row2 = tk.Frame(win)
        btn_row2.pack(fill="x", padx=fs(10), pady=(0, fs(10)))
        tk.Button(
            btn_row2, text="🔧 เติม HN ให้คนที่ยังไม่มี", font=("Tahoma", fs(9)),
            command=assign_missing,
        ).pack(side="left")

        refresh()
        win.lift()
        win.focus_force()

    def open_ai_assist_dialog(self):
        """Phase 1 of an "AI ช่วยค้นข้อมูล" helper - free-text only (อาการ/
        อายุ/เพศ typed in by the pharmacist), nothing sourced from or linked
        to any patient record in this app (no name, no phone, no document/
        photo upload - deliberately, to keep this a general lookup tool and
        not a channel for real patient data). Reference-only: never treat
        the reply as a diagnosis - that's why the disclaimer is pinned at
        the top, not buried in a tooltip somewhere."""
        settings = app_settings.load_settings()
        configured = [k for k in ai_assist.PROVIDERS if settings.get(ai_assist.PROVIDERS[k]["key_field"])]

        win = tk.Toplevel(self.root)
        win.title("AI ช่วยค้นข้อมูล")
        win.geometry(f"{fs(520)}x{fs(600)}")
        win.transient(self.root)
        win.grab_set()

        tk.Label(
            win, text="⚠ คำตอบจาก AI เป็นข้อมูลอ้างอิงประกอบการตัดสินใจเท่านั้น ไม่ใช่คำวินิจฉัยทางการแพทย์ "
                       "- เภสัชกรต้องใช้วิจารณญาณตัดสินใจเองเสมอ ห้ามพิมพ์ชื่อหรือข้อมูลส่วนตัวของผู้ป่วยลงในนี้",
            font=("Tahoma", fs(9), "bold"), fg="#a00", wraplength=fs(490), justify="left",
        ).pack(anchor="w", padx=fs(10), pady=(fs(10), fs(6)))

        if not configured:
            tk.Label(
                win, text="ยังไม่ได้ตั้งค่า API Key เลย - ไปที่ปุ่ม ⚙️ ตั้งค่า เพื่อใส่ API Key ของ ChatGPT/Claude/Grok "
                           "อย่างน้อย 1 ตัวก่อนใช้งานฟีเจอร์นี้",
                font=("Tahoma", fs(10), "bold"), fg="#a00", wraplength=fs(490), justify="left",
            ).pack(anchor="w", padx=fs(10), pady=(0, fs(10)))
            tk.Button(win, text="ปิด", font=("Tahoma", fs(10)), command=win.destroy).pack(pady=fs(10))
            win.lift()
            win.focus_force()
            return

        provider_row = tk.Frame(win)
        provider_row.pack(fill="x", padx=fs(10), pady=(0, fs(6)))
        tk.Label(provider_row, text="ผู้ให้บริการ AI:", font=("Tahoma", fs(10), "bold")).pack(side="left")
        provider_labels = [ai_assist.PROVIDERS[k]["label"] for k in configured]
        provider_display_var = tk.StringVar(value=provider_labels[0])
        provider_combo = ttk.Combobox(
            provider_row, textvariable=provider_display_var, values=provider_labels,
            state="readonly", font=("Tahoma", fs(10)),
        )
        provider_combo.pack(side="left", padx=(fs(6), 0), fill="x", expand=True)
        label_to_key = {ai_assist.PROVIDERS[k]["label"]: k for k in configured}

        tk.Label(
            win, text="พิมพ์อาการ/อายุ/เพศ (ห้ามใส่ชื่อหรือข้อมูลที่ระบุตัวตนได้)",
            font=("Tahoma", fs(10), "bold"),
        ).pack(anchor="w", padx=fs(10))
        prompt_text = tk.Text(win, font=("Tahoma", fs(11)), height=6, bd=1, relief="solid", wrap="word")
        prompt_text.pack(fill="x", padx=fs(10), pady=(fs(2), fs(6)))

        status_var = tk.StringVar(value="")
        tk.Label(win, textvariable=status_var, font=("Tahoma", fs(9)), fg="#666").pack(anchor="w", padx=fs(10))
        source_var = tk.StringVar(value="")
        tk.Label(win, textvariable=source_var, font=("Tahoma", fs(8)), fg="#1a7a4a",
                 wraplength=fs(490), justify="left").pack(anchor="w", padx=fs(10))

        # Kept only for this dialog's lifetime, never persisted. Sent back
        # on every follow-up so the AI has context of what was already
        # asked/answered - this is what makes it "remember", at the cost of
        # growing token usage per turn (each new call resends the whole
        # running conversation, not just the latest question).
        history = []  # [{"role": "user"/"assistant", "text": ...}, ...]

        # Packed here (right after the prompt box) rather than at the very
        # end - a tall response Text with expand=True would otherwise push
        # these buttons below the bottom of the window on shorter screens,
        # same overflow problem seen earlier with the edit-drug dialog.
        btn_row = tk.Frame(win)
        btn_row.pack(pady=(fs(2), fs(8)))
        send_btn = tk.Button(
            btn_row, text="📤 ส่ง", font=("Tahoma", fs(10), "bold"), bg="#1a5a9a", fg="white", command=lambda: do_send(),
        )
        send_btn.pack(side="left", padx=(0, fs(6)))
        clear_btn = tk.Button(
            btn_row, text="🗑 ล้าง (เริ่มเรื่องใหม่)", font=("Tahoma", fs(10)), command=lambda: do_clear(),
        )
        clear_btn.pack(side="left")

        tk.Label(win, text="บทสนทนา", font=("Tahoma", fs(10), "bold")).pack(anchor="w", padx=fs(10), pady=(fs(4), 0))
        response_text = tk.Text(win, font=("Tahoma", fs(10)), bd=1, relief="solid", wrap="word", state="disabled")
        response_text.pack(fill="both", expand=True, padx=fs(10), pady=(fs(2), fs(8)))

        def append_transcript(label, text):
            response_text.config(state="normal")
            if response_text.index("1.0") != response_text.index("end-1c"):
                response_text.insert(tk.END, "\n\n")
            response_text.insert(tk.END, f"{label}\n{text}")
            response_text.config(state="disabled")
            response_text.see(tk.END)

        def do_clear():
            history.clear()
            prompt_text.delete("1.0", tk.END)
            response_text.config(state="normal")
            response_text.delete("1.0", tk.END)
            response_text.config(state="disabled")
            status_var.set("")
            source_var.set("")
            prompt_text.focus_set()

        def do_send():
            prompt = prompt_text.get("1.0", tk.END).strip()
            if not prompt:
                status_var.set("กรุณาพิมพ์อาการก่อน")
                return
            provider_key = label_to_key.get(provider_display_var.get())
            if not provider_key:
                return
            provider_info = ai_assist.PROVIDERS[provider_key]
            api_key = settings.get(provider_info["key_field"], "")
            send_btn.config(state="disabled")
            status_var.set(f"กำลังส่งไปที่ {provider_info['label']}...")
            source_var.set("")
            prompt_text.delete("1.0", tk.END)
            append_transcript("🧑 เภสัชกร", prompt)

            def worker():
                results = knowledge.search_knowledge(prompt)
                context = knowledge.format_context_block(results)
                history_block = ""
                if history:
                    lines = [f"{'เภสัชกรถาม' if h['role'] == 'user' else 'AI ตอบ'}: {h['text']}" for h in history]
                    history_block = "ประวัติการสนทนาก่อนหน้าในหัวข้อนี้ (ใช้ประกอบบริบท):\n" + "\n".join(lines) + "\n\n"
                full_prompt = (
                    f"{context}\n\n{history_block}คำถามล่าสุดจากเภสัชกร:\n{prompt}"
                    if (context or history_block) else prompt
                )
                success, text = provider_info["call"](api_key, full_prompt)
                def apply():
                    send_btn.config(state="normal")
                    status_var.set("" if success else "เกิดข้อผิดพลาด")
                    if results:
                        names = ", ".join(sorted({r["source"] for r in results}))
                        source_var.set(f"📎 อ้างอิงจากเอกสาร: {names}")
                    else:
                        source_var.set("ℹ️ ไม่พบข้อมูลที่เกี่ยวข้องในเอกสารอ้างอิง (ตอบจากความรู้ทั่วไปของ AI)")
                    append_transcript("🤖 AI", text)
                    if success:
                        history.append({"role": "user", "text": prompt})
                        history.append({"role": "assistant", "text": text})
                self.root.after(0, apply)

            threading.Thread(target=worker, daemon=True).start()

        win.lift()
        win.focus_force()
        prompt_text.focus_set()

    # ---------------------------------------------------------------- confirm + print

    def on_confirm(self):
        if not self.selected_drugs:
            messagebox.showwarning("แจ้งเตือน", "กรุณาเลือกยาอย่างน้อย 1 รายการ")
            return
        self.open_patient_dialog()

    def open_patient_dialog(self):
        total_labels = sum(d.get("print_qty", 1) for d in self.selected_drugs)

        win = tk.Toplevel(self.root)
        win.title("ชื่อผู้ป่วย")
        win.geometry(f"{fs(360)}x{fs(300)}")
        win.transient(self.root)
        win.grab_set()

        tk.Label(win, text=f"พร้อมพิมพ์ {total_labels} ฉลาก ({len(self.selected_drugs)} รายการยา)",
                 font=("Tahoma", fs(10))).pack(pady=(fs(14), fs(4)))
        tk.Label(win, text="ชื่อ นามสกุล (ไม่บังคับ)", font=("Tahoma", fs(10), "bold")).pack(anchor="w", padx=fs(14))
        # Not cleared here on open (only in on_print_done, once actually
        # printed) - closing/cancelling this dialog to go fix something in
        # the drug list shouldn't lose the picked name; reopening it should
        # still show it.
        patient_var = tk.StringVar(value=self._queue_patient_name or "")
        entry = tk.Entry(win, textvariable=patient_var, font=("Tahoma", fs(12)))
        entry.pack(fill="x", padx=fs(14), pady=fs(4))
        entry.select_range(0, "end")

        anon_var = tk.BooleanVar(value=False)

        def on_anon_toggle():
            if anon_var.get():
                entry.config(state="disabled")
            else:
                entry.config(state="normal")
                entry.focus_set()

        tk.Checkbutton(
            win, text="ไม่ประสงค์ออกนาม", variable=anon_var, font=("Tahoma", fs(10)),
            command=on_anon_toggle,
        ).pack(anchor="w", padx=fs(14))

        # Saving to a patient file is the uncommon case (most prints don't
        # need it) - kept out of the default view entirely. The phone field
        # and name/phone search-autocomplete only exist inside this popup,
        # opened on demand, so the common no-save path never touches the DB
        # with a search query.
        phone_var = tk.StringVar(value=self._queue_patient_phone or "")  # also only cleared in on_print_done
        save_state = {"save": False}
        save_status_var = tk.StringVar(value="")

        save_row = tk.Frame(win)
        save_row.pack(fill="x", padx=fs(14), pady=(fs(10), 0))
        tk.Button(
            save_row, text="💾 บันทึกลงแฟ้มผู้ป่วย", font=("Tahoma", fs(9), "bold"), fg="#1a5a9a",
            command=lambda: open_save_popup(),
        ).pack(side="left")

        tk.Label(
            win, textvariable=save_status_var, font=("Tahoma", fs(9), "bold"), fg="#1a7a4a",
            wraplength=fs(320), justify="left",
        ).pack(anchor="w", padx=fs(14), pady=(fs(2), 0))

        def open_save_popup():
            popup = tk.Toplevel(win)
            popup.title("บันทึกลงแฟ้มผู้ป่วย")
            popup.geometry(f"{fs(360)}x{fs(360)}")
            popup.transient(win)
            popup.grab_set()

            tk.Label(popup, text="ชื่อ นามสกุล", font=("Tahoma", fs(10), "bold")).pack(anchor="w", padx=fs(14), pady=(fs(12), 0))
            p_name_var = tk.StringVar(value=patient_var.get().strip())
            p_name_entry = tk.Entry(popup, textvariable=p_name_var, font=("Tahoma", fs(12)))
            p_name_entry.pack(fill="x", padx=fs(14), pady=fs(4))

            # Autocomplete: suggest existing patients as the pharmacist
            # types, so a returning patient's name+phone can be picked in
            # one click instead of retyped (retyping is the #1 way
            # name-matching to patient history silently breaks - a typo
            # means it never links up later).
            name_suggest = tk.Listbox(popup, font=("Tahoma", fs(9)), height=3, exportselection=False)
            name_suggest.pack(fill="x", padx=fs(14))
            name_suggest_results = []

            def on_name_typed(*_a):
                term = p_name_var.get().strip()
                name_suggest.delete(0, tk.END)
                name_suggest_results.clear()
                if not term:
                    return
                try:
                    name_suggest_results.extend(storage.search_patients(term, limit=6))
                except Exception:
                    return
                for p in name_suggest_results:
                    phone_part = f" - {p['phone']}" if p["phone"] else ""
                    name_suggest.insert(tk.END, f"{p['name']}{phone_part}")

            def on_name_pick(event=None):
                sel = name_suggest.curselection()
                if not sel:
                    return
                p = name_suggest_results[sel[0]]
                p_name_var.set(p["name"])
                p_phone_var.set(p["phone"])
                name_suggest.delete(0, tk.END)
                phone_suggest.delete(0, tk.END)

            p_name_var.trace_add("write", on_name_typed)
            name_suggest.bind("<<ListboxSelect>>", on_name_pick)

            tk.Label(popup, text="เบอร์โทร", font=("Tahoma", fs(10), "bold")).pack(anchor="w", padx=fs(14), pady=(fs(6), 0))
            p_phone_var = tk.StringVar(value=phone_var.get().strip())
            tk.Entry(popup, textvariable=p_phone_var, font=("Tahoma", fs(12))).pack(fill="x", padx=fs(14), pady=fs(4))

            phone_suggest = tk.Listbox(popup, font=("Tahoma", fs(9)), height=3, exportselection=False)
            phone_suggest.pack(fill="x", padx=fs(14))
            phone_suggest_results = []

            def on_phone_typed(*_a):
                term = p_phone_var.get().strip()
                phone_suggest.delete(0, tk.END)
                phone_suggest_results.clear()
                if not term:
                    return
                try:
                    phone_suggest_results.extend(storage.search_patients(term, limit=6))
                except Exception:
                    return
                for p in phone_suggest_results:
                    phone_part = f" - {p['phone']}" if p["phone"] else ""
                    phone_suggest.insert(tk.END, f"{p['name']}{phone_part}")

            def on_phone_pick(event=None):
                sel = phone_suggest.curselection()
                if not sel:
                    return
                p = phone_suggest_results[sel[0]]
                p_name_var.set(p["name"])
                p_phone_var.set(p["phone"])
                phone_suggest.delete(0, tk.END)
                name_suggest.delete(0, tk.END)

            p_phone_var.trace_add("write", on_phone_typed)
            phone_suggest.bind("<<ListboxSelect>>", on_phone_pick)

            def confirm_save():
                p_name = p_name_var.get().strip()
                p_phone = p_phone_var.get().strip()
                if not p_name and not p_phone:
                    messagebox.showwarning("แจ้งเตือน", "กรุณาใส่ชื่อหรือเบอร์โทรก่อนบันทึก", parent=popup)
                    return
                if p_name and not p_phone:
                    try:
                        existing = storage.find_patients_by_exact_name(p_name)
                    except Exception:
                        existing = []
                    if existing:
                        messagebox.showwarning(
                            "ชื่อนี้มีอยู่แล้ว",
                            f"มีผู้ป่วยชื่อ \"{p_name}\" อยู่ในระบบแล้ว {len(existing)} คน "
                            "กรุณาระบุเบอร์โทรเพื่อยืนยันว่าเป็นคนเดียวกัน (หรือคนละคน) ก่อนบันทึก",
                            parent=popup,
                        )
                        return
                patient_var.set(p_name)
                phone_var.set(p_phone)
                save_state["save"] = True
                save_status_var.set("✓ จะบันทึกลงแฟ้ม: " + p_name + (f" - {p_phone}" if p_phone else ""))
                popup.destroy()

            btn_row = tk.Frame(popup)
            btn_row.pack(pady=(fs(10), fs(10)))
            tk.Button(
                btn_row, text="ยืนยันบันทึก", font=("Tahoma", fs(10), "bold"), bg="#1a7a4a", fg="white",
                command=confirm_save,
            ).pack(side="left", padx=fs(4))
            tk.Button(btn_row, text="ยกเลิก", font=("Tahoma", fs(10)), command=popup.destroy).pack(side="left", padx=fs(4))

            popup.lift()
            popup.focus_force()
            p_name_entry.focus_set()

        status_var = tk.StringVar(value="")
        tk.Label(win, textvariable=status_var, font=("Tahoma", fs(9)), fg="#070", wraplength=fs(320)).pack(pady=(fs(4), 0))

        def do_print():
            # both name and phone are optional now - printing an unnamed
            # label is a valid choice, not an error the user has to work
            # around by ticking "anonymous" every single time
            name = "ไม่ประสงค์ออกนาม" if anon_var.get() else patient_var.get().strip()
            phone = phone_var.get().strip()
            print_btn.config(state="disabled")
            status_var.set("กำลังพิมพ์...")

            def worker():
                try:
                    settings = app_settings.load_settings()
                    # patient_id only ever comes from a real patients-table
                    # record - either just created/found here (save ticked),
                    # or already resolved unambiguously by pick_name(). A
                    # typed name with no save and no prior pick stays
                    # unlinked (None), same as before this feature existed.
                    if save_state["save"]:
                        patient_id = storage.find_or_create_patient(name, phone)
                    else:
                        patient_id = self._queue_patient_id
                    has_allergy = self.allergy_var.get()
                    allergy_drug_name = self._get_allergy_drug_name(patient_id) if has_allergy else ""
                    # Build every physical label first (one PIL image per
                    # copy, print_qty duplicates reuse the same rendered
                    # image object - content is identical per copy) before
                    # deciding how to send them to the printer, since A4
                    # mode needs to see the whole batch to tile them 2-up
                    # per sheet instead of printing one-per-page.
                    label_imgs = []
                    for d in self.selected_drugs:
                        data = dict(d)
                        data["patient_name"] = name
                        data["has_allergy"] = has_allergy
                        data["allergy_drug_name"] = allergy_drug_name
                        lang = d.get("lang", "th")
                        note = (d.get("note") or "").strip()
                        if lang in ("en", "mm") and note:
                            # Already threaded (this whole worker runs off
                            # the UI thread) - safe to call Grok
                            # synchronously here. Reuses the cache from
                            # preview_label() if the pharmacist already
                            # previewed; translates on the spot otherwise,
                            # so printing without ever clicking Preview
                            # still produces a translated Indication line
                            # instead of silently printing Thai text.
                            translated = d.get(f"note_{lang}")
                            if not translated:
                                translated, _err = translate_note_via_grok(note, lang)
                                if translated:
                                    d[f"note_{lang}"] = translated
                                    idproduct = d.get("idproduct")
                                    if idproduct:
                                        storage.save_note_translation(idproduct, lang, translated)
                            if translated:
                                data["note"] = translated
                        img = build_label_image(data, settings)
                        label_imgs.extend([img] * d.get("print_qty", 1))
                    if label_imgs:
                        label_imgs[-1].save(DEBUG_PREVIEW_PATH)
                    if settings.get("paper_mode") == "a4":
                        for page in build_a4_pages(label_imgs):
                            print_image(page)
                    else:
                        for img in label_imgs:
                            print_image(img)
                    storage.add_print_job(name, phone, self.selected_drugs, patient_id=patient_id)
                    self.root.after(0, lambda: self.on_print_done(win, total_labels))
                except Exception as e:
                    self.root.after(0, lambda: status_var.set(f"เกิดข้อผิดพลาด: {e}"))
                    self.root.after(0, lambda: print_btn.config(state="normal"))

            threading.Thread(target=worker, daemon=True).start()

        def save_history_only():
            # For when the pharmacist just wants the visit remembered for
            # next time (so it shows up in "แฟ้มประวัติการจ่ายยา" /
            # "เลือกชื่อลูกค้า" later) without actually sending anything to
            # the printer - e.g. a phone consult or a "what did they buy
            # last time" lookup that didn't end in a real dispense.
            name = "ไม่ประสงค์ออกนาม" if anon_var.get() else patient_var.get().strip()
            phone = phone_var.get().strip()
            history_btn.config(state="disabled")
            print_btn.config(state="disabled")
            status_var.set("กำลังบันทึกประวัติ...")

            def worker():
                try:
                    if save_state["save"]:
                        patient_id = storage.find_or_create_patient(name, phone)
                    else:
                        patient_id = self._queue_patient_id
                    storage.add_print_job(name, phone, self.selected_drugs, patient_id=patient_id)
                    self.root.after(0, lambda: self.on_history_saved(win))
                except Exception as e:
                    self.root.after(0, lambda: status_var.set(f"เกิดข้อผิดพลาด: {e}"))
                    self.root.after(0, lambda: history_btn.config(state="normal"))
                    self.root.after(0, lambda: print_btn.config(state="normal"))

            threading.Thread(target=worker, daemon=True).start()

        btn_row2 = tk.Frame(win)
        btn_row2.pack(pady=fs(10))
        print_btn = tk.Button(btn_row2, text="\U0001F5A8️ พิมพ์ฉลาก", font=("Tahoma", fs(11), "bold"),
                               bg="#1a7a4a", fg="white", command=do_print)
        print_btn.pack(side="left", padx=fs(4))
        history_btn = tk.Button(
            btn_row2, text="📋 บันทึกประวัติฉลาก (ไม่พิมพ์)", font=("Tahoma", fs(9)),
            command=save_history_only,
        )
        history_btn.pack(side="left", padx=fs(4))

        win.bind("<Return>", lambda e: do_print())
        entry.focus_set()
        win.lift()
        win.focus_force()

    def on_print_done(self, win, total_labels):
        win.destroy()
        self.status_var.set(f"พิมพ์ครบ {total_labels} ฉลากแล้ว")
        self.selected_drugs = []
        self.allergy_var.set(False)
        self.selected_customer_var.set(self.NO_CUSTOMER_TEXT)
        self._queue_patient_name = None
        self._queue_patient_phone = None
        self._queue_patient_id = None
        self.refresh_selected_list()

    def on_history_saved(self, win):
        win.destroy()
        self.status_var.set("บันทึกประวัติแล้ว (ไม่ได้พิมพ์ฉลาก)")
        self.selected_drugs = []
        self.allergy_var.set(False)
        self.selected_customer_var.set(self.NO_CUSTOMER_TEXT)
        self._queue_patient_name = None
        self._queue_patient_phone = None
        self._queue_patient_id = None
        self.refresh_selected_list()


def main():
    root = tk.Tk()
    # NOTE: do NOT withdraw() root before showing the first-run settings
    # Toplevel. On Windows, a Toplevel with transient(parent) set can fail to
    # ever become visible if parent is withdrawn at the time transient() is
    # called - the window manager ties the child's showability to the
    # (invisible) owner. Keep root visible (even if blank for a moment)
    # until after the dialog closes.

    if not app_settings.settings_exist():
        setup_win = build_settings_dialog(root, first_run=True)
        root.wait_window(setup_win)

    root.deiconify()
    LabelApp(root)
    root.lift()
    root.attributes("-topmost", True)
    root.after(200, lambda: root.attributes("-topmost", False))
    root.focus_force()
    root.mainloop()


if __name__ == "__main__":
    main()
